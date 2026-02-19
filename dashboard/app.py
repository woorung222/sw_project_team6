import streamlit as st
import pandas as pd
import psycopg2
import plotly.express as px
import json
import subprocess
import os
import glob
import time
import io
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re


# 1. 초기 설정
st.set_page_config(page_title="Security Audit System", page_icon="🛡️", layout="wide")

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');

    /* =========================
       Base (Theme-safe)
       - 다크모드 깨짐 방지: background/color 강제 지정 제거
       - 대신 카드/보더만 theme-friendly 하게
       ========================= */
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

    /* App container background (Streamlit theme에 맡김) */
    [data-testid="stAppViewContainer"] { background: transparent; }
    [data-testid="stHeader"] { background: transparent; }

    /* Header */
    .main-header {
        background: transparent;
        padding: 14px 24px;
        border-bottom: 1px solid rgba(128,128,128,0.22);
        margin-bottom: 16px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 12px;
    }
    .main-header h3 { font-size: 20px; margin: 0; font-weight: 700; }

    .header-right {
        display: flex;
        align-items: center;
        gap: 10px;
        white-space: nowrap;
    }
    .header-user {
        font-size: 12px;
        opacity: 0.75;
    }

    /* KPI card (theme-safe) - menu button vibe */
.kpi-card {
    padding: 18px;
    border-radius: 14px;
    border: 1px solid rgba(128,128,128,0.28);
    border-left: 4px solid;
    margin-bottom: 10px;
    backdrop-filter: blur(6px);
    box-shadow: 0 6px 18px rgba(0,0,0,0.12);
    background: rgba(255,255,255,0.05);
}
@media (prefers-color-scheme: light) {
    .kpi-card { background: rgba(0,0,0,0.03); box-shadow: 0 6px 18px rgba(0,0,0,0.06); }
}
.kpi-title { font-size: 12px; font-weight: 700; opacity: 0.8; }
.kpi-value { font-size: 26px; font-weight: 800; margin: 6px 0 0 0; }

[data-testid="stMetricValue"] { font-size: 22px !important; }
    [data-testid="stMetricLabel"] { font-size: 14px !important; }

  /* =========================
   Button Style (NO RED)
   - Streamlit 기본 primary(빨강) 완전 제거
   - Sidebar는 네비 버튼 느낌 유지
   ========================= */

/* Streamlit theme primary color 변수도 함께 오버라이드(버전에 따라 적용) */
:root {
    --primary-color: #0EA5E9;
}

/* ---- Sidebar: nav buttons (high contrast) ---- */
[data-testid="stSidebar"] .stButton > button,
[data-testid="stSidebar"] [data-testid="stDownloadButton"] button,
[data-testid="stSidebar"] [data-testid="stFormSubmitButton"] button {
    background: rgba(255,255,255,0.02) !important;
    color: inherit !important;
    border: 1px solid rgba(255,255,255,0.20) !important;
    border-radius: 10px !important;
    padding: 10px 12px !important;
    text-align: left !important;
    font-weight: 700 !important;
    letter-spacing: 0.1px !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] .stButton > button:hover,
[data-testid="stSidebar"] [data-testid="stDownloadButton"] button:hover,
[data-testid="stSidebar"] [data-testid="stFormSubmitButton"] button:hover {
    background: rgba(59,130,246,0.10) !important;
    border-color: rgba(59,130,246,0.40) !important;
    filter: none !important;
}
[data-testid="stSidebar"] .stButton > button:focus-visible,
[data-testid="stSidebar"] [data-testid="stDownloadButton"] button:focus-visible,
[data-testid="stSidebar"] [data-testid="stFormSubmitButton"] button:focus-visible {
    outline: 2px solid rgba(255,255,255,0.55) !important;
    outline-offset: 2px !important;
}

/* ---- Main area: ALL action buttons unified (no red) ---- */
[data-testid="stAppViewContainer"] section.main .stButton > button,
[data-testid="stAppViewContainer"] section.main [data-testid="stDownloadButton"] button,
[data-testid="stAppViewContainer"] section.main [data-testid="stFormSubmitButton"] button {
    background: linear-gradient(180deg, rgba(14,165,233,1) 0%, rgba(2,132,199,1) 100%) !important;
    color: #ffffff !important;
    border: 1px solid rgba(2,132,199,0.55) !important;
    border-radius: 10px !important;
    padding: 10px 14px !important;
    font-weight: 800 !important;
    box-shadow: 0 10px 24px rgba(2,132,199,0.18) !important;
    transition: transform 0.06s ease-in-out, filter 0.15s ease-in-out !important;
}

/* Hover/active */
[data-testid="stAppViewContainer"] section.main .stButton > button:hover,
[data-testid="stAppViewContainer"] section.main [data-testid="stDownloadButton"] button:hover,
[data-testid="stAppViewContainer"] section.main [data-testid="stFormSubmitButton"] button:hover {
    filter: brightness(0.95) !important;
}
[data-testid="stAppViewContainer"] section.main .stButton > button:active,
[data-testid="stAppViewContainer"] section.main [data-testid="stDownloadButton"] button:active,
[data-testid="stAppViewContainer"] section.main [data-testid="stFormSubmitButton"] button:active {
    transform: translateY(1px) !important;
    filter: brightness(0.90) !important;
}

/* Disabled */
[data-testid="stAppViewContainer"] section.main .stButton > button:disabled,
[data-testid="stAppViewContainer"] section.main [data-testid="stDownloadButton"] button:disabled,
[data-testid="stAppViewContainer"] section.main [data-testid="stFormSubmitButton"] button:disabled {
    background: rgba(148,163,184,0.35) !important;
    border-color: rgba(148,163,184,0.35) !important;
    color: rgba(255,255,255,0.75) !important;
    box-shadow: none !important;
    cursor: not-allowed !important;
}

/* ---- Extra safety: some Streamlit versions render primary as st-emotion-cache classes.
   Catch buttons that still end up with the default red by forcing any main button background away from red hues. ---- */
[data-testid="stAppViewContainer"] section.main button {
    accent-color: #0EA5E9 !important;
}

/* Download button (Report) - keep same palette */
  section.main div.stDownloadButton > button,
  section.main div.stDownloadButton button {
      background-color: #334155 !important;
      color: #ffffff !important;
      border: none !important;
      border-radius: 8px !important;
      transition: filter 0.15s ease-in-out !important;
  }
  section.main div.stDownloadButton > button:hover,
  section.main div.stDownloadButton button:hover {
      filter: brightness(0.92) !important;
  }

  /* Focus outline / weird red glow 방지 */
  section.main .stButton > button:focus,
  section.main .stButton > button:focus-visible,
  section.main button[kind="primary"]:focus,
  section.main button[kind="primary"]:focus-visible {
      outline: none !important;
      box-shadow: none !important;
  }


    /* Sidebar */
    [data-testid="stSidebar"] { border-right: 1px solid rgba(128,128,128,0.22); }
    .nav-section { font-size: 12px; font-weight: 700; opacity: 0.75; margin: 10px 0 6px 0; }
    .nav-btn > div { padding: 0 !important; } /* 버튼 간격 통일 */

    /* Placeholder / Logs */
    .guide-placeholder {
        background: transparent;
        border: 1px dashed rgba(128,128,128,0.35);
        padding: 15px;
        border-radius: 8px;
        opacity: 0.85;
        font-size: 13px;
    }

    .execution-log {
        background-color: #1e1e1e;
        color: #d4d4d4;
        padding: 15px;
        border-radius: 8px;
        font-family: 'Courier New', monospace;
        font-size: 12px;
        border: 1px solid rgba(255,255,255,0.10);
        height: 350px;
        overflow-y: scroll;
    }

    div[data-testid="column"] button { margin-top: 28px; }

    .result-card {
        background-color: #262730 !important; /* KPI 카드와 동일한 색상 */
        border: 1px solid rgba(255, 255, 255, 0.1) !important; /* 은은한 외곽선 */
        border-radius: 12px;
        padding: 18px;
        margin: 12px 0;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.3); /* 깊이감을 위한 그림자 */
        color: #ffffff !important;
    }

    .result-title { 
        font-weight: 700; 
        font-size: 15px; 
        margin-bottom: 8px; 
        color: #F0F2F6 !important; /* 제목 강조 */
    }

    .muted { 
        color: rgba(255, 255, 255, 0.6) !important; /* 가독성을 고려한 밝은 회색 */
        font-size: 12px; 
        line-height: 1.5;
    }

    /* 카드 내부에 구분선이 필요할 경우 사용 */
    .result-card hr {
        border: 0;
        border-top: 1px solid rgba(255, 255, 255, 0.05);
        margin: 10px 0;
    }

    /* =========================
       Grid width normalize
       - Selected 컬럼 과도한 폭 방지
       ========================= */
    /* 데이터에디터 첫번째 컬럼(Selected) 폭 고정 (Streamlit 내부 구조 변동 대비 약하게 적용) */
    [data-testid="stDataFrame"] div[role="columnheader"]:first-child,
    [data-testid="stDataFrame"] div[role="gridcell"]:first-child {
        min-width: 72px !important;
        max-width: 72px !important;
        width: 72px !important;
    }

    /* =========================
        Dark mode overrides
        - OS/브라우저 다크 선호 + Streamlit theme dark 둘 다 커버
        ========================= */
    @media (prefers-color-scheme: dark), html[data-theme="dark"] {
        .kpi-card, .result-card {
            background-color: #000000 !important; /* Streamlit 기본 배경보다 약간 밝은 회색 */
            border: 1px solid rgba(255,255,255,0.1) !important;
            box-shadow: 0 4px 12px rgba(0,0,0,0.4) !important; /* 깊이감 있는 그림자 */
            color: #ffffff !important;
        }

        /* 카드 내부 텍스트 색상 강제 지정 */
        .kpi-title { color: rgba(255,255,255,0.7) !important; }
        .kpi-value { color: #ffffff !important; }
        .result-title { color: #ffffff !important; }
        .muted { color: rgba(255,255,255,0.5) !important; }

        /* 경계선 및 가이드라인 조정 */
        [data-testid="stSidebar"] { border-right: 1px solid rgba(255,255,255,0.1) !important; }
        .main-header { border-bottom: 1px solid rgba(255,255,255,0.1) !important; }
        .guide-placeholder { 
            border: 1px dashed rgba(255,255,255,0.2) !important; 
            background-color: rgba(255,255,255,0.02) !important;
        }
    }
    </style>
""",
    unsafe_allow_html=True,
)

# 2. 세션 상태 관리
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "user_info" not in st.session_state:
    st.session_state.user_info = {}
if "current_page" not in st.session_state:
    st.session_state.current_page = "Dashboard"
if "os_selection" not in st.session_state:
    st.session_state.os_selection = []
if "db_selection" not in st.session_state:
    st.session_state.db_selection = []
if "rep_host_select" not in st.session_state:
    st.session_state.rep_host_select = []
if "execution_output" not in st.session_state:
    st.session_state.execution_output = ""

# 조치 결과/리프레시 관련
if "last_remedy_result" not in st.session_state:
    st.session_state.last_remedy_result = None
if "needs_refresh_after_remedy" not in st.session_state:
    st.session_state.needs_refresh_after_remedy = False

# Start Audit 완료 요약
if "last_audit_summary" not in st.session_state:
    st.session_state.last_audit_summary = None

# Batch Remedy 결과 요약(새 결과창)
if "last_batch_result" not in st.session_state:
    st.session_state.last_batch_result = None


# =========================
# Report hard-coded config
# =========================

UNIX_KEYWORDS = [
    "ubuntu", "rocky", "rhel", "red hat", "centos", "debian", "suse", "opensuse",
    "amazon linux", "almalinux", "oracle linux", "aix", "solaris", "hp-ux", "unix", "linux"
]

CATEGORY_MAP = {
    "account": "계정 관리",
    "file": "파일 및 디렉터리 관리",
    "service": "서비스 관리",
    "patch": "패치 관리",
    "log": "로그",
}

CATEGORY_ORDER = [
    "계정 관리",
    "파일 및 디렉터리 관리",
    "서비스 관리",
    "패치 관리",
    "로그",
]


# 3. 헬퍼 함수

def fetch_latest_scans_for_hosts(hostnames: list[str]) -> pd.DataFrame:
    """
    선택된 호스트들에 대해 host+flag 기준 최신 1건을 가져온다.
    - 최신 기준: round DESC -> scan_at DESC -> scan_id DESC
    반환 컬럼:
      hostname, ip_address, os_version, flag_id, category, risk_level, description,
      is_vul, scan_at, msg
    """
    if not hostnames:
        return pd.DataFrame()

    conn = get_db_connection()
    query = """
        WITH latest AS (
            SELECT
                s.*,
                ROW_NUMBER() OVER (
                    PARTITION BY s.host_id, s.flag_id
                    ORDER BY s.round DESC, s.scan_at DESC, s.scan_id DESC
                ) AS rn
            FROM audit.scans s
        )
        SELECT
            h.hostname,
            h.ip_address,
            h.os_version,
            l.flag_id,
            g.category,
            g.risk_level,
            g.description,
            CASE WHEN l.is_vul = TRUE THEN 1 ELSE 0 END AS is_vul,
            l.scan_at,
            l.msg,
            l.round
        FROM latest l
        JOIN audit.hosts h ON h.host_id = l.host_id
        LEFT JOIN audit.guides g ON g.flag_id = l.flag_id
        WHERE l.rn = 1
          AND h.hostname = ANY(%s);
    """
    df = pd.read_sql(query, conn, params=(hostnames,))
    conn.close()

    # 방어
    if df.empty:
        return df

    df["flag_id"] = df["flag_id"].fillna("N/A").astype(str)
    df["category"] = df["category"].fillna("미분류").astype(str)
    df["risk_level"] = df["risk_level"].fillna("Low").astype(str)
    df["description"] = df["description"].fillna("").astype(str)
    df["msg"] = df["msg"].fillna("").astype(str)

    return df


def _risk_to_letter(risk_level: str) -> str:
    s = str(risk_level).lower()
    if "high" in s or "상" in s:
        return "H"
    if "med" in s or "중" in s:
        return "M"
    return "L"


def _flag_to_report_code(parent_flag: str) -> str:
    # 그룹형 그대로: U_07 -> U-07
    return str(parent_flag).replace("_", "-")


def _sort_parent_flags(flags: list[str]) -> list[str]:
    # U_07 같은 형태에서 숫자 기준 정렬
    def key_fn(x: str):
        m = re.search(r"U[_-](\d+)", str(x))
        return int(m.group(1)) if m else 999999
    return sorted(flags, key=key_fn)


def _normalize_category(cat_raw: str) -> str:
    s = str(cat_raw).strip().lower()
    # 흔한 변형 처리
    if s.startswith("account"):
        return CATEGORY_MAP["account"]
    if s.startswith("file"):
        return CATEGORY_MAP["file"]
    if s.startswith("service"):
        return CATEGORY_MAP["service"]
    if s.startswith("patch"):
        return CATEGORY_MAP["patch"]
    if s.startswith("log"):
        return CATEGORY_MAP["log"]
    # 혹시 이미 한글이면 그대로
    for v in CATEGORY_MAP.values():
        if v in s:
            return v
    return "로그" if "log" in s else "서비스 관리" if "service" in s else "파일 및 디렉터리 관리" if "file" in s else "계정 관리" if "account" in s else "패치 관리" if "patch" in s else "계정 관리"


def _is_unix_os(os_version: str) -> bool:
    s = str(os_version or "").lower()
    return any(k in s for k in UNIX_KEYWORDS)


def _build_parent_meta_fixed67() -> pd.DataFrame:
    """
    U_01~U_67 고정 + DB guide 정보로 메타 구성.
    - guides에 없으면 placeholder
    - category는 account/file/service/patch/log 를 한글로 매핑
    """
    wanted = [f"U_{i:02d}" for i in range(1, 68)]

    conn = get_db_connection()
    guides_df = pd.read_sql(
        """
        SELECT flag_id, category, risk_level, description
        FROM audit.guides
        WHERE flag_id LIKE 'U_%'
        """,
        conn,
    )
    conn.close()

    if guides_df.empty:
        # 전부 placeholder
        rows = []
        for p in wanted:
            idx = int(re.search(r"U_(\d+)", p).group(1)) if re.search(r"U_(\d+)", p) else 0
            rows.append(
                {
                    "parent_flag": p,
                    "category": "계정 관리",
                    "description": str(idx),  # placeholder
                    "risk_level": "Low",
                }
            )
        return pd.DataFrame(rows)

    guides_df["flag_id"] = guides_df["flag_id"].astype(str)
    guides_df["parent_flag"] = guides_df["flag_id"].apply(get_parent_flag)
    guides_df["category_norm"] = guides_df["category"].apply(_normalize_category)

    def _risk_rank(x: str) -> int:
        s = str(x).lower()
        if "high" in s or "상" in s:
            return 3
        if "med" in s or "중" in s:
            return 2
        return 1

    # parent 기준 대표 정보
    parent_meta = (
        guides_df.groupby("parent_flag", as_index=False)
        .agg(
            category=("category_norm", lambda s: next((v for v in s if str(v).strip()), "계정 관리")),
            description=("description", lambda s: next((v for v in s if str(v).strip()), "")),
            risk_level=("risk_level", lambda s: sorted(list(s), key=_risk_rank, reverse=True)[0] if len(s) else "Low"),
        )
    )

    # fixed 67 rows
    meta_map = {r["parent_flag"]: r for _, r in parent_meta.iterrows()}
    out_rows = []
    for p in wanted:
        idx = int(re.search(r"U_(\d+)", p).group(1)) if re.search(r"U_(\d+)", p) else 0
        if p in meta_map:
            desc = str(meta_map[p].get("description") or "").strip()
            if not desc:
                desc = str(idx)  # placeholder
            out_rows.append(
                {
                    "parent_flag": p,
                    "category": meta_map[p].get("category", "계정 관리"),
                    "description": desc,
                    "risk_level": meta_map[p].get("risk_level", "Low"),
                }
            )
        else:
            out_rows.append(
                {
                    "parent_flag": p,
                    "category": "계정 관리",
                    "description": str(idx),  # placeholder
                    "risk_level": "Low",
                }
            )

    df = pd.DataFrame(out_rows)

    # category order 정렬 + parent numeric 정렬
    df["cat_order"] = df["category"].apply(lambda x: CATEGORY_ORDER.index(x) if x in CATEGORY_ORDER else 999)
    df["num"] = df["parent_flag"].apply(lambda x: int(re.search(r"U_(\d+)", x).group(1)) if re.search(r"U_(\d+)", x) else 999)
    df = df.sort_values(["cat_order", "num"]).drop(columns=["cat_order", "num"]).reset_index(drop=True)
    return df


def _excel_style_base(ws):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return border


def _set_col_widths(ws, widths: dict[int, float]):
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def generate_unix_report_excel() -> io.BytesIO:
    """
    [보고서 정책 - 최종 확정 반영]
    - UNIX 계열만
    - 항목 1~67 고정 (U_01 ~ U_67 하드코딩)
    - 점검 내용(description)은 DB에서 가져오지 않음: "U_NN 점검 내용" 하드코딩
    - 카테고리 매핑은 기존 코드(CATEGORY_MAP/CATEGORY_ORDER) 유지
    - 조치 결과 열: 공백
    - 취약(현재 설정): scan 로그(.log)에서 basis 추출 (최신 ts 1개만 사용)
    - 양호(조치 내용 및 미조치 사유): remediate 로그(.log)에서 result 추출 (최신 ts 1개만 사용)
    - 점수 열: 공백
    - 색상 정책: 진단결과가 '취약'이면 글자 빨강 + 셀 배경 연빨강
    - B안: scan/remedy 연결 안함(run_id 미사용), 각각 최신 1개만 사용
    - 로그는 .log만 사용(.txt 무시)
    """

    # 1) 전체 호스트 가져오기 + UNIX 필터
    conn = get_db_connection()
    hosts_df = pd.read_sql(
        """
        SELECT hostname, ip_address, os_version
        FROM audit.hosts
        WHERE hostname IS NOT NULL
        ORDER BY hostname;
        """,
        conn,
    )
    conn.close()

    if hosts_df.empty:
        out = io.BytesIO()
        out.write(b"")
        out.seek(0)
        return out

    hosts_df["hostname"] = hosts_df["hostname"].astype(str)
    hosts_df["ip_address"] = hosts_df["ip_address"].fillna("").astype(str)
    hosts_df["os_version"] = hosts_df["os_version"].fillna("").astype(str)

    hosts_df = hosts_df[hosts_df["os_version"].apply(_is_unix_os)].copy().reset_index(drop=True)

    if hosts_df.empty:
        out = io.BytesIO()
        out.write(b"")
        out.seek(0)
        return out

    hosts_df["host_code"] = hosts_df.index.map(lambda i: f"U{(i+1):03d}")

    # 2) 스캔/조치 최신값 준비 (B안 유지: scan/remedy 연결 안함)
    hostnames = hosts_df["hostname"].tolist()

    # (1) audit.scans: host + parent_flag(U_01~U_67) 기준 최신 1건
    scan_parent_latest = pd.DataFrame()
    try:
        conn = get_db_connection()
        scans_all = pd.read_sql(
            """
            SELECT
                h.hostname,
                s.scan_id,
                s.flag_id,
                s.is_vul,
                s.msg,
                s.round,
                s.scan_at
            FROM audit.scans s
            JOIN audit.hosts h ON h.host_id = s.host_id
            WHERE h.hostname = ANY(%s)
              AND s.flag_id LIKE 'U%'
            """,
            conn,
            params=(hostnames,),
        )
        conn.close()

        if not scans_all.empty:
            scans_all["flag_id"] = scans_all["flag_id"].astype(str)
            scans_all["parent_flag"] = scans_all["flag_id"].apply(get_parent_flag)
            scans_all["msg"] = scans_all["msg"].fillna("").astype(str)
            scans_all["round"] = pd.to_numeric(scans_all["round"], errors="coerce").fillna(-1)
            scans_all["scan_at"] = pd.to_datetime(scans_all["scan_at"], errors="coerce")
            scans_all["scan_id"] = pd.to_numeric(scans_all["scan_id"], errors="coerce").fillna(-1)

            scans_all = scans_all.sort_values(["hostname", "parent_flag", "round", "scan_at", "scan_id"], ascending=[True, True, False, False, False])
            scan_parent_latest = scans_all.drop_duplicates(["hostname", "parent_flag"], keep="first").reset_index(drop=True)
    except Exception:
        scan_parent_latest = pd.DataFrame()

    scan_parent_map: dict[tuple[str, str], dict] = {}
    if not scan_parent_latest.empty:
        for _, r in scan_parent_latest.iterrows():
            scan_parent_map[(str(r["hostname"]), str(r["parent_flag"]))] = {
                "scan_id": r.get("scan_id"),
                "flag_id": str(r.get("flag_id") or ""),
                "is_vul": bool(r.get("is_vul")) if pd.notna(r.get("is_vul")) else None,
                "msg": str(r.get("msg") or ""),
                "round": int(r.get("round")) if pd.notna(r.get("round")) else None,
                "scan_at": r.get("scan_at"),
            }

    # (2) audit.remedy: host + parent_flag(U_01~U_67) 기준 최신 1건
    remedy_parent_latest = pd.DataFrame()
    try:
        conn = get_db_connection()
        remedy_all = pd.read_sql(
            """
            SELECT
                h.hostname,
                r.remedy_id,
                r.flag_id,
                r.description,
                r.status,
                r.scan_at
            FROM audit.remedy r
            JOIN audit.hosts h ON h.host_id = r.host_id
            WHERE h.hostname = ANY(%s)
              AND r.flag_id LIKE 'U%'
            """,
            conn,
            params=(hostnames,),
        )
        conn.close()

        if not remedy_all.empty:
            remedy_all["flag_id"] = remedy_all["flag_id"].astype(str)
            remedy_all["parent_flag"] = remedy_all["flag_id"].apply(get_parent_flag)
            remedy_all["description"] = remedy_all["description"].fillna("").astype(str)
            remedy_all["scan_at"] = pd.to_datetime(remedy_all["scan_at"], errors="coerce")
            remedy_all["remedy_id"] = pd.to_numeric(remedy_all["remedy_id"], errors="coerce").fillna(-1)

            remedy_all = remedy_all.sort_values(["hostname", "parent_flag", "scan_at", "remedy_id"], ascending=[True, True, False, False])
            remedy_parent_latest = remedy_all.drop_duplicates(["hostname", "parent_flag"], keep="first").reset_index(drop=True)
    except Exception:
        remedy_parent_latest = pd.DataFrame()

    remedy_parent_map: dict[tuple[str, str], dict] = {}
    if not remedy_parent_latest.empty:
        for _, r in remedy_parent_latest.iterrows():
            remedy_parent_map[(str(r["hostname"]), str(r["parent_flag"]))] = {
                "remedy_id": r.get("remedy_id"),
                "flag_id": str(r.get("flag_id") or ""),
                "description": str(r.get("description") or ""),
                "status": r.get("status"),
                "scan_at": r.get("scan_at"),
            }

    parent_meta = _build_parent_meta_fixed67()

    # -----------------------------
    # 로그 파싱 (직접 파싱)
    # - scan: status/basis (section == "[진단 결과]") 최신 ts 1개
    # - remediate: result 최신 ts 1개
    # - .log만 사용(.txt 무시)
    # -----------------------------
    def _parse_ts(ts_raw: str):
        if not ts_raw:
            return None
        s = str(ts_raw).strip()

        # "2026-02-18T21:11:28+0900" -> "+09:00"
        m_kr = re.match(r"^(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})([+-]\d{2})(\d{2})$", s)
        if m_kr:
            s = f"{m_kr.group(1)}{m_kr.group(2)}:{m_kr.group(3)}"

        try:
            if s.endswith("Z"):
                return datetime.fromisoformat(s.replace("Z", "+00:00"))
            return datetime.fromisoformat(s)
        except Exception:
            return None

    def _iter_json_objects_from_log(fpath: str):
        """.log에서 JSON 객체들을 안전하게 추출한다.
        - JSONL(한 줄 1객체) / pretty-printed JSON(여러 줄 1객체) / [..] 배열 모두 지원
        - 정책: .log만 사용, 실패하면 빈 iterator
        """
        if not fpath or (not isinstance(fpath, str)) or (not fpath.endswith(".log")) or (not os.path.exists(fpath)):
            return
        try:
            with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                raw = f.read()
            raw_strip = raw.strip()
            if not raw_strip:
                return

            # 1) 파일 전체가 단일 JSON(객체/배열)인 경우
            try:
                j = json.loads(raw_strip)
                if isinstance(j, dict):
                    yield j
                    return
                if isinstance(j, list):
                    for it in j:
                        if isinstance(it, dict):
                            yield it
                    return
            except Exception:
                pass

            # 2) JSONL: 줄 단위 시도
            yielded_any = False
            for line in raw.splitlines():
                s = line.strip()
                if not s or not s.startswith("{"):
                    continue
                try:
                    obj = json.loads(s)
                    if isinstance(obj, dict):
                        yielded_any = True
                        yield obj
                except Exception:
                    continue
            if yielded_any:
                return

            # 3) pretty-printed 객체 추출(중괄호 카운팅)
            buf = []
            depth = 0
            in_str = False
            esc = False
            for ch in raw:
                buf.append(ch)
                if in_str:
                    if esc:
                        esc = False
                        continue
                    if ch == "\\":
                        esc = True
                        continue
                    if ch == '"':
                        in_str = False
                    continue
                else:
                    if ch == '"':
                        in_str = True
                        continue
                    if ch == "{":
                        depth += 1
                    elif ch == "}":
                        depth -= 1
                        if depth == 0:
                            chunk = "".join(buf).strip()
                            buf = []
                            try:
                                obj = json.loads(chunk)
                                if isinstance(obj, dict):
                                    yield obj
                            except Exception:
                                continue
            return
        except Exception:
            return

    def _extract_u_num(flag_like: str):
        m_u = re.search(r"U[_-]?(\d+)", str(flag_like))
        if not m_u:
            return None
        try:
            return int(m_u.group(1))
        except Exception:
            return None

    def _candidate_log_dirs(kind: str) -> list[str]:
        # kind: "scan" | "remediate"
        return [
            f"/home/audit/isms-ansible/results/{kind}/log",
            os.path.join(os.getcwd(), "results", kind, "log"),
            os.path.join(os.getcwd(), "results", kind, "log", "archive"),
        ]

    def _safe_is_log_file(path: str) -> bool:
        # 정확히 .log 로 끝나는 것만 인정(.log.txt 등은 무시)
        return bool(path) and isinstance(path, str) and path.endswith(".log") and os.path.exists(path)

    def _find_latest_log_file(kind: str, hostname: str, parent_flag: str) -> str | None:
        """
        1) scan은 DB msg가 실제 .log 경로면 그걸 우선 사용(정확히 .log만)
        2) 아니면 표준 results 경로에서 glob로 탐색
        """
        u_num = _extract_u_num(parent_flag)
        if u_num is None:
            return None

        # 1) msg 경로 우선 (scan만)
        if kind == "scan":
            msg_path = scan_msg_map.get((hostname, parent_flag), "")
            if _safe_is_log_file(msg_path):
                return msg_path

        # 2) glob 탐색
        patterns = [
            f"*{hostname}*{kind}*U_{u_num:02d}.log",
            f"*{hostname}*{kind}*U_{u_num}.log",
            f"*{hostname}*U_{u_num:02d}.log",
            f"*{hostname}*U_{u_num}.log",
            f"*{kind}*U_{u_num:02d}.log",
            f"*{kind}*U_{u_num}.log",
        ]

        candidates: list[str] = []
        for base_dir in _candidate_log_dirs(kind):
            if not os.path.isdir(base_dir):
                continue
            for pat in patterns:
                candidates.extend(glob.glob(os.path.join(base_dir, pat)))

        candidates = [p for p in candidates if p.endswith(".log")]

        if not candidates:
            return None

        try:
            return sorted(candidates, key=os.path.getmtime, reverse=True)[0]
        except Exception:
            return candidates[0]

    # host별 OS tag(로그 파일명 prefix 추정) 맵
    host_os_map: dict[str, str] = {str(r["hostname"]): str(r["os_version"]) for _, r in hosts_df.iterrows()}

    def _guess_os_tag(os_version: str) -> str | None:
        s = str(os_version or "").lower()
        # 파일 트리 기준: ubuntu24 / rocky9
        if "ubuntu" in s:
            return "ubuntu24" if ("24" in s or "24.04" in s) else "ubuntu24"
        if "rocky" in s:
            return "rocky9" if ("9" in s) else "rocky9"
        if "centos" in s:
            return "rocky9"
        if "rhel" in s or "redhat" in s:
            return "rocky9"
        return None

    def _extract_log_path_from_text(text: str) -> str | None:
        """msg/description에 로그 경로가 섞여 들어오는 케이스 방어"""
        if not text:
            return None
        # 가장 흔한 절대경로 .log 추출
        m = re.search(r"(\/[^\s\"]+\.log)", str(text))
        if m:
            p = m.group(1)
            if os.path.exists(p) and p.endswith(".log"):
                return p
        # 혹시 상대경로라면 그대로 시도
        if str(text).endswith(".log") and os.path.exists(str(text)):
            return str(text)
        return None

    def _find_latest_log_by_os(kind: str, os_tag: str | None, parent_flag: str) -> str | None:
        """hostname이 파일명에 없어서, OS tag + U_NN으로만 찾는 fallback (정확도 낮음)"""
        if not os_tag:
            return None
        u_num = _extract_u_num(parent_flag)
        if u_num is None:
            return None
        patterns = [
            f"{os_tag}-{kind}_U_{u_num:02d}.log",
            f"{os_tag}-{kind}_U_{u_num}.log",
            f"{os_tag}-{kind}_U_{u_num:02d}.log.*",  # 혹시 확장자 변형
        ]
        candidates: list[str] = []
        for base_dir in _candidate_log_dirs(kind):
            if not os.path.isdir(base_dir):
                continue
            for pat in patterns:
                candidates.extend(glob.glob(os.path.join(base_dir, pat)))
        # 정책: .log만 사용(.txt, .log.txt 등 무시)
        candidates = [p for p in candidates if p.endswith(".log")]
        if not candidates:
            return None
        try:
            return sorted(candidates, key=os.path.getmtime, reverse=True)[0]
        except Exception:
            return candidates[0]

    def _read_latest_scan_entry(hostname: str, parent_flag: str) -> dict:
        """
        scan 로그(.log)에서:
        - DB의 최신 scan row에서 msg(로그 경로)를 우선 사용 (host별 로그 매칭 핵심)
        - msg로 로그를 못 찾는 경우에만 OS tag 기반 fallback
        - 최신 ts 1개만 사용
        - 최신 ts 구간에서:
          * section == "[진단 내용]"  : title/cmd/result 모두 수집
          * section == "[진단 결과]" : status/basis 모두 수집 (동일 ts에 여러 결과가 있으면 합침)
        """
        # 1) DB latest row의 msg에서 로그 경로 확보
        fpath = None
        row = scan_parent_map.get((str(hostname), str(parent_flag)))
        if row:
            fpath = _extract_log_path_from_text(row.get("msg", ""))

        # 2) fallback: OS tag 기반 (동일 OS 다중 호스트 환경에서는 오염 가능)
        if not fpath:
            os_tag = _guess_os_tag(host_os_map.get(str(hostname), ""))
            fpath = _find_latest_log_by_os("scan", os_tag, parent_flag)

        if not fpath:
            return {}

        want_num = _extract_u_num(parent_flag)

        # 1차: 최신 ts(dt) 찾기 (진단 내용/진단 결과 모두 포함)
        best_dt = None
        best_ts_raw = None
        try:
            with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    line = line.strip()
                    if not line.startswith("{"):
                        continue
                    try:
                        obj = json.loads(line)
                    except Exception:
                        continue

                    num = _extract_u_num(obj.get("flag_id"))
                    if want_num is not None and num != want_num:
                        continue

                    section = str(obj.get("section") or "").strip()
                    if section not in ("[진단 내용]", "[진단 결과]"):
                        continue

                    dt = _parse_ts(obj.get("ts"))
                    if dt is None:
                        continue

                    if best_dt is None or dt > best_dt:
                        best_dt = dt
                        best_ts_raw = obj.get("ts")
        except Exception:
            return {}

        if best_dt is None:
            return {}

        # 2차: 최신 ts 구간 내용 수집
        diag_items = []   # [진단 내용]
        result_items = [] # [진단 결과]
        try:
            with open(fpath, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    line = line.strip()
                    if not line.startswith("{"):
                        continue
                    try:
                        obj = json.loads(line)
                    except Exception:
                        continue

                    num = _extract_u_num(obj.get("flag_id"))
                    if want_num is not None and num != want_num:
                        continue

                    section = str(obj.get("section") or "").strip()
                    if section not in ("[진단 내용]", "[진단 결과]"):
                        continue

                    dt = _parse_ts(obj.get("ts"))
                    if dt is None or dt != best_dt:
                        continue

                    if section == "[진단 내용]":
                        title = str(obj.get("title") or "").strip()
                        cmd = str(obj.get("cmd") or "").strip()
                        result = str(obj.get("result") or "").strip()
                        # 최소 1개 필드라도 있으면 수집
                        if title or cmd or result:
                            diag_items.append({"title": title, "cmd": cmd, "result": result})

                    elif section == "[진단 결과]":
                        status = str(obj.get("status") or "").strip()
                        basis = str(obj.get("basis") or "").strip()
                        if status or basis:
                            result_items.append({"status": status, "basis": basis})
        except Exception:
            return {}

        # status 종합 (하나라도 취약이면 취약, 아니면 양호)
        statuses = [it["status"] for it in result_items if it.get("status")]
        if any("취약" in s for s in statuses):
            status_final = "취약"
        elif any("양호" in s for s in statuses):
            status_final = "양호"
        else:
            status_final = statuses[0] if statuses else ""

        # basis 합치기 (중복 제거)
        bases = []
        seen = set()
        for it in result_items:
            b = str(it.get("basis") or "").strip()
            if not b:
                continue
            if b in seen:
                continue
            seen.add(b)
            bases.append(b)
        basis_final = "\n".join(bases).strip()

        # 진단 내용 포맷 (중복 제거: title+cmd+result 기준)
        details_lines = []
        details_seen = set()
        for it in diag_items:
            title = (it.get("title") or "").strip()
            cmd = (it.get("cmd") or "").strip()
            result = (it.get("result") or "").strip()
            key = f"{title}||{cmd}||{result}"
            if key in details_seen:
                continue
            details_seen.add(key)

            if title:
                details_lines.append(f"{title}")
            if cmd:
                details_lines.append(f"cmd: {cmd}")
            if result:
                details_lines.append(f"result: {result}")
            # 항목 구분 줄
            details_lines.append("")

        details_final = "\n".join(details_lines).strip()

        return {
            "ts": best_ts_raw,
            "status": status_final,
            "basis": basis_final,
            "details": details_final,
            "source_file": fpath,
        }

    def _read_latest_remedy_entry(hostname: str, parent_flag: str) -> dict:
        """
        remediate 로그(.log)에서:
        - 최신 ts 묶음 1개만 사용
        - section == "[조치 결과]"의 title/cmd/result 전부 수집
        - 사람이 읽기 좋은 텍스트로 합쳐 반환
        - '조치 판정'은 DB(audit.remedy) 값을 쓰지 않고, remediate.log에서만 판단
          * 최신 ts 묶음에서 result에 "조치 성공" 또는 "status: 2" 포함, 또는 status==2 이면 조치 성공
        """
        # 1) 로그 파일 찾기: DB description에 경로가 있으면 우선(경로로만 사용), 아니면 OS tag 기반
        fpath = None
        row = remedy_parent_map.get((str(hostname), str(parent_flag)))
        if row:
            fpath = _extract_log_path_from_text(row.get("description", ""))

        if not fpath:
            os_tag = _guess_os_tag(host_os_map.get(str(hostname), ""))
            fpath = _find_latest_log_by_os("remediate", os_tag, parent_flag)

        if not fpath:
            return {}

        want_num = _extract_u_num(parent_flag)

        # 최신 ts 선택
        best_dt = None
        best_ts_raw = None
        try:
            for obj in _iter_json_objects_from_log(fpath):
                num = _extract_u_num(obj.get("flag_id"))
                if want_num is not None and num != want_num:
                    continue
                dt = _parse_ts(obj.get("ts"))
                if dt is None:
                    continue
                if best_dt is None or dt > best_dt:
                    best_dt = dt
                    best_ts_raw = obj.get("ts")
        except Exception:
            best_dt = None
            best_ts_raw = None

        if best_dt is None:
            return {}

        # 최신 ts 묶음에서 [조치 결과] 수집
        items: list[dict] = []
        try:
            for obj in _iter_json_objects_from_log(fpath):
                num = _extract_u_num(obj.get("flag_id"))
                if want_num is not None and num != want_num:
                    continue
                dt = _parse_ts(obj.get("ts"))
                if dt is None or dt != best_dt:
                    continue
                if str(obj.get("section") or "").strip() != "[조치 결과]":
                    continue
                items.append(obj)
        except Exception:
            items = []

        # 사람이 읽기 좋게 합치기
        lines = []
        seen = set()
        for it in items:
            title = str(it.get("title") or "").strip()
            cmd = str(it.get("cmd") or "").strip()
            result = str(it.get("result") or "").strip()
            key = f"{title}||{cmd}||{result}"
            if key in seen:
                continue
            seen.add(key)
            if title:
                lines.append(title)
            if cmd:
                lines.append(f"cmd: {cmd}")
            if result:
                lines.append(f"result: {result}")
            lines.append("")

        details_text = "\n".join(lines).strip()


        # 조치 성공 판정(최신 ts 묶음 내부)
        success = False
        try:
            for it in items:
                r = str(it.get("result") or "")
                if "조치 성공" in r:
                    success = True
                    break
                if "status: 2" in r:
                    success = True
                    break
                stv = it.get("status")
                try:
                    if int(stv) == 2:
                        success = True
                        break
                except Exception:
                    pass
        except Exception:
            success = False

        return {"ts": best_ts_raw, "success": bool(success), "details": details_text, "source_file": fpath}
    scan_log_cache: dict[tuple[str, str], dict] = {}
    remedy_log_cache: dict[tuple[str, str], dict] = {}

    def _get_scan_log(hostname: str, pflag: str) -> dict:
        key = (hostname, pflag)
        if key not in scan_log_cache:
            scan_log_cache[key] = _read_latest_scan_entry(hostname, pflag)
        return scan_log_cache[key]

    def _get_remedy_log(hostname: str, pflag: str) -> dict:
        key = (hostname, pflag)
        if key not in remedy_log_cache:
            remedy_log_cache[key] = _read_latest_remedy_entry(hostname, pflag)
        return remedy_log_cache[key]

    def _u_check_content(pflag: str) -> str:
        # 점검 내용 하드코딩 (상/중/하 표시는 무시)
        U_CONTENT_MAP = {
            1: "root 계정 원격 접속 제한",
            2: "비밀번호 관리정책 설정",
            3: "계정 잠금 임계값 설정",
            4: "비밀번호 파일 보호",
            5: "root 이외의 UID가 ‘0’ 금지",
            6: "사용자 계정 su 기능 제한",
            7: "불필요한 계정 제거",
            8: "관리자 그룹에 최소한의 계정 포함",
            9: "계정이 존재하지 않는 GID 금지",
            10: "동일한 UID 금지",
            11: "사용자 Shell 점검",
            12: "세션 종료 시간 설정",
            13: "안전한 비밀번호 암호화 알고리즘 사용",
            14: "root 홈, 패스 디렉터리 권한 및 패스 설정",
            15: "파일 및 디렉터리 소유자 설정",
            16: "/etc/passwd 파일 소유자 및 권한 설정",
            17: "시스템 시작 스크립트 권한 설정",
            18: "/etc/shadow 파일 소유자 및 권한 설정",
            19: "/etc/hosts 파일 소유자 및 권한 설정",
            20: "/etc/(x)inetd.conf 파일 소유자 및 권한 설정",
            21: "/etc/(r)syslog.conf 파일 소유자 및 권한 설정",
            22: "/etc/services 파일 소유자 및 권한 설정",
            23: "SUID, SGID, Sticky bit 설정 파일 점검",
            24: "사용자, 시스템 환경변수 파일 소유자 및 권한 설정",
            25: "world writable 파일 점검",
            26: "/dev에 존재하지 않는 device 파일 점검",
            27: "$HOME/.rhosts, hosts.equiv 사용 금지",
            28: "접속 IP 및 포트 제한",
            29: "hosts.lpd 파일 소유자 및 권한 설정",
            30: "UMASK 설정 관리",
            31: "홈 디렉토리 소유자 및 권한 설정",
            32: "홈 디렉토리로 지정한 디렉터리의 존재 관리",
            33: "숨겨진 파일 및 디렉터리 검색 및 제거",
            34: "Finger 서비스 비활성화",
            35: "공유 서비스에 대한 익명 접근 제한 설정",
            36: "r 계열 서비스 비활성화",
            37: "crontab 설정파일 권한 설정 미흡",
            38: "DoS 공격에 취약한 서비스 비활성화",
            39: "불필요한 NFS 서비스 비활성화",
            40: "NFS 접근 통제",
            41: "불필요한 automountd 제거",
            42: "불필요한 RPC 서비스 비활성화",
            43: "NIS, NIS+ 점검",
            44: "tftp, talk 서비스 비활성화",
            45: "메일 서비스 버전 점검",
            46: "일반 사용자의 메일 서비스 실행 방지",
            47: "스팸 메일 릴레이 제한",
            48: "expn, vrfy 명령어 제한",
            49: "DNS 보안 버전 패치",
            50: "DNS Zone Transfer 설정",
            51: "DNS 서비스의 취약한 동적 업데이트 설정 금지",
            52: "Telnet 서비스 비활성화",
            53: "FTP 서비스 정보 노출 제한",
            54: "암호화되지 않는 FTP 서비스 비활성화",
            55: "FTP 계정 Shell 제한",
            56: "FTP 서비스 접근 제어 설정",
            57: "Ftpusers 파일 설정",
            58: "불필요한 SNMP 서비스 구동 점검",
            59: "안전한 SNMP 버전 사용",
            60: "SNMP Community String 복잡성 설정",
            61: "SNMP Access Control 설정",
            62: "로그인 시 경고 메시지 설정",
            63: "sudo 명령어 접근 관리",
            64: "주기적 보안 패치 및 벤더 권고사항 적용",
            65: "NTP 및 시각 동기화 설정",
            66: "정책에 따른 시스템 로깅 설정",
            67: "로그 디렉터리 소유자 및 권한 설정",
        }

        m_u = re.search(r"U_(\d+)", str(pflag))
        num = int(m_u.group(1)) if m_u else 0
        return U_CONTENT_MAP.get(num, "")

    def get_result_text(hostname: str, pflag: str) -> str:
        # 1) DB(is_vul) 우선 (host+parent_flag 최신 1건)
        key = (str(hostname), str(pflag))
        row = scan_parent_map.get(key)
        if row and row.get("is_vul") is not None:
            return "취약" if bool(row.get("is_vul")) else "양호"

        # 2) DB에 없으면 scan 로그(status/basis)로 fallback
        info = _get_scan_log(hostname, pflag)
        if not info:
            return "N/A"
        status = str(info.get("status") or "").strip()
        if "취약" in status:
            return "취약"
        if "양호" in status:
            return "양호"
        basis = str(info.get("basis") or "").strip()
        if basis:
            return "취약"
        return "N/A"
    def get_scan_basis(hostname: str, pflag: str) -> str:
        """scan 로그에서 최신 ts 구간의 진단 내용(cmd/result) + 진단 결과(basis/status)를 합쳐 반환"""
        info = _get_scan_log(hostname, pflag)
        if not info:
            return ""

        details = str(info.get("details") or "").strip()
        basis = str(info.get("basis") or "").strip()
        status = str(info.get("status") or "").strip()

        parts = []
        if details:
            parts.append("[진단 내용]")
            parts.append(details)
        if basis or status:
            parts.append("[진단 결과]")
            if status:
                parts.append(f"status: {status}")
            if basis:
                parts.append(basis)

        return "\n".join(parts).strip()

    def get_remedy_result(hostname: str, pflag: str) -> str:
        info = _get_remedy_log(hostname, pflag)
        return str(info.get("details") or "").strip() if info else ""
    def get_remedy_success(hostname: str, pflag: str) -> bool:
        info = _get_remedy_log(hostname, pflag)
        return bool(info.get("success")) if info else False


    # 4) workbook 생성 (기존 구조 유지)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    border = _excel_style_base(None)
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    title_fill = PatternFill("solid", fgColor="D9E1F2")
    vuln_fill = PatternFill("solid", fgColor="F8D7DA")  # 연빨
    vuln_font = Font(color="C00000", bold=True)
    normal_font = Font(color="000000")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 5) 표지
    ws_cover = wb.create_sheet("표지")
    ws_cover.merge_cells("A1:H1")
    ws_cover["A1"].value = "서버 취약점진단 상세결과(UNIX)"
    ws_cover["A1"].font = Font(bold=True, size=18)
    ws_cover["A1"].alignment = center

    now = datetime.now()
    date_str = f"{now.year}년 {now.month}월 {now.day}일"
    ws_cover.merge_cells("A3:H3")
    ws_cover["A3"].value = f"작성일: {date_str}"
    ws_cover["A3"].alignment = center
    ws_cover["A3"].font = Font(size=12)

    for r in range(1, 30):
        ws_cover.row_dimensions[r].height = 22
    _set_col_widths(ws_cover, {1: 20, 2: 20, 3: 20, 4: 20, 5: 20, 6: 20, 7: 20, 8: 20})

    # 6) 서버진단대상
    ws_targets = wb.create_sheet("서버진단대상")
    headers = ["OS", "NO.", "Hostname", "IP", "OS Version", "비고"]
    ws_targets.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws_targets.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for idx, r in hosts_df.iterrows():
        ws_targets.append(
            [
                "OS",
                r["host_code"],
                r["hostname"],
                r["ip_address"],
                r["os_version"],
                "",  # 비고
            ]
        )

    for r in range(2, ws_targets.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws_targets.cell(r, c)
            cell.border = border
            cell.alignment = left if c in [3, 4, 5, 6] else center

    _set_col_widths(ws_targets, {1: 8, 2: 10, 3: 22, 4: 18, 5: 28, 6: 14})
    ws_targets.freeze_panes = "A2"

    # 7) UNIX_요약
    ws_sum = wb.create_sheet("UNIX_요약")

    # 타이틀
    ws_sum.merge_cells("A1:D1")
    ws_sum["A1"].value = "UNIX 점검 요약"
    ws_sum["A1"].fill = title_fill
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A1"].alignment = center

    # 헤더
    base_headers = ["영역", "점검항목", "점검내용", "위험도"]
    host_headers = hosts_df["host_code"].tolist()
    all_headers = base_headers + host_headers
    ws_sum.append(all_headers)

    for c in range(1, len(all_headers) + 1):
        cell = ws_sum.cell(row=2, column=c)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # 본문
    row_ptr = 3
    last_cat = None
    for _, it in parent_meta.iterrows():
        cat = it["category"]
        pflag = it["parent_flag"]
        desc = _u_check_content(pflag)
        risk_letter = _risk_to_letter(it["risk_level"])

        ws_sum.cell(row_ptr, 1).value = cat if cat != last_cat else ""
        ws_sum.cell(row_ptr, 2).value = _flag_to_report_code(pflag)
        ws_sum.cell(row_ptr, 3).value = desc
        ws_sum.cell(row_ptr, 4).value = risk_letter

        for hi, hrow in hosts_df.iterrows():
            hostname = hrow["hostname"]
            res = get_result_text(hostname, pflag)
            cell = ws_sum.cell(row_ptr, 5 + hi)
            remedy_success = get_remedy_success(hostname, pflag)
            display_res = res
            if res == "양호" and remedy_success:
                display_res = "조치"
            cell.value = display_res
            if display_res == "취약":
                cell.fill = vuln_fill
                cell.font = vuln_font
            else:
                cell.font = normal_font

        # style
        for c in range(1, 5 + len(host_headers)):
            cell = ws_sum.cell(row_ptr, c)
            cell.border = border
            cell.alignment = left if c == 3 else center

        last_cat = cat
        row_ptr += 1

    _set_col_widths(ws_sum, {1: 18, 2: 10, 3: 52, 4: 8})
    for i in range(len(host_headers)):
        _set_col_widths(ws_sum, {5 + i: 10})
    ws_sum.freeze_panes = "A3"

    # 8) 통계 및 그래프 (표만)
    ws_stat = wb.create_sheet("통계 및 그래프")

    ws_stat.merge_cells("A1:F1")
    ws_stat["A1"].value = "통계 및 그래프 (표)"
    ws_stat["A1"].fill = title_fill
    ws_stat["A1"].font = Font(bold=True, size=14)
    ws_stat["A1"].alignment = center

    # 집계
    total_good = 0
    total_vul = 0
    total_na = 0

    cat_map = {c: [0, 0, 0] for c in CATEGORY_ORDER}  # good, vul, na
    risk_map = {"H": [0, 0, 0], "M": [0, 0, 0], "L": [0, 0, 0]}

    for _, it in parent_meta.iterrows():
        cat = it["category"]
        pflag = it["parent_flag"]
        rletter = _risk_to_letter(it["risk_level"])
        if cat not in cat_map:
            cat_map[cat] = [0, 0, 0]

        for _, hrow in hosts_df.iterrows():
            hostname = hrow["hostname"]
            res = get_result_text(hostname, pflag)

            if res == "취약":
                total_vul += 1
                cat_map[cat][1] += 1
                risk_map[rletter][1] += 1
            elif res == "양호":
                total_good += 1
                cat_map[cat][0] += 1
                risk_map[rletter][0] += 1
            else:
                total_na += 1
                cat_map[cat][2] += 1
                risk_map[rletter][2] += 1

    denom = (total_good + total_vul)
    vuln_rate = (total_vul / denom * 100.0) if denom > 0 else 0.0

    def write_table(start_row: int, title: str, headers_: list[str], rows_: list[list]):
        ws_stat.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers_))
        tcell = ws_stat.cell(start_row, 1)
        tcell.value = title
        tcell.fill = header_fill
        tcell.font = Font(bold=True)
        tcell.alignment = left

        for c, h in enumerate(headers_, start=1):
            cell = ws_stat.cell(start_row + 1, c)
            cell.value = h
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        r = start_row + 2
        for row in rows_:
            for c, v in enumerate(row, start=1):
                cell = ws_stat.cell(r, c)
                cell.value = v
                cell.border = border
                cell.alignment = center
            r += 1

        return r

    # 표 1: 전체 통계
    rows_total = [
        ["전체", total_good, total_vul, total_na, f"{vuln_rate:.1f}%"],
    ]
    r_next = write_table(
        3,
        "전체 통계",
        ["구분", "양호", "취약", "N/A", "취약률(%)"],
        rows_total,
    )

    # 표 2: 영역별
    rows_cat = []
    for c in CATEGORY_ORDER:
        g, v, na = cat_map.get(c, [0, 0, 0])
        denom2 = g + v
        rate2 = (v / denom2 * 100.0) if denom2 > 0 else 0.0
        rows_cat.append([c, g, v, na, f"{rate2:.1f}%"])

    r_next = write_table(
        r_next + 1,
        "영역별 점검 결과",
        ["영역", "양호", "취약", "N/A", "취약률(%)"],
        rows_cat,
    )

    # 표 3: 위험도별
    rows_risk = []
    for rl in ["H", "M", "L"]:
        g, v, n = risk_map[rl]
        denom3 = g + v
        vr = (v / denom3 * 100.0) if denom3 > 0 else 0.0
        rows_risk.append([rl, int(g), int(v), int(n), round(vr, 2)])

    _ = write_table(
        r_next + 1,
        "위험도별 위험 수준",
        ["위험도", "양호", "취약", "N/A", "취약률(%)"],
        rows_risk,
    )

    _set_col_widths(ws_stat, {1: 22, 2: 16, 3: 12, 4: 12, 5: 12, 6: 14})

    # 10) 개별 시트 (U001..)
    indiv_headers = ["영역", "점검항목", "점검내용", "위험도", "진단결과", "조치결과", "취약(현재 설정)", "양호(조치 내용 및 미조치시 사유)", "점수"]

    for idx, hrow in hosts_df.iterrows():
        code = hrow["host_code"]
        hostname = hrow["hostname"]
        ip = hrow["ip_address"]
        os_ver = hrow["os_version"]

        ws_host = wb.create_sheet(code)

        ws_host.merge_cells("A1:I1")
        ws_host["A1"].value = f"개별 점검 결과 - {code}"
        ws_host["A1"].fill = title_fill
        ws_host["A1"].font = Font(bold=True, size=13)
        ws_host["A1"].alignment = center

        ws_host["A2"].value = "Hostname"
        ws_host["B2"].value = hostname
        ws_host["A3"].value = "IP"
        ws_host["B3"].value = ip
        ws_host["D2"].value = "OS"
        ws_host["E2"].value = "OS"
        ws_host["D3"].value = "OS Version"
        ws_host["E3"].value = os_ver

        for cell_addr in ["A2", "A3", "D2", "D3"]:
            ws_host[cell_addr].font = Font(bold=True)
        for addr in ["A2", "A3", "B2", "B3", "D2", "D3", "E2", "E3"]:
            ws_host[addr].alignment = left

        start_row = 6
        for c, h in enumerate(indiv_headers, start=1):
            cell = ws_host.cell(start_row, c)
            cell.value = h
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        rptr = start_row + 1
        last_cat = None
        for _, it in parent_meta.iterrows():
            cat = it["category"]
            pflag = it["parent_flag"]
            desc = _u_check_content(pflag)
            risk_letter = _risk_to_letter(it["risk_level"])
            res = get_result_text(hostname, pflag)
        
            ws_host.cell(rptr, 1).value = cat if cat != last_cat else ""
            ws_host.cell(rptr, 2).value = _flag_to_report_code(pflag)
            ws_host.cell(rptr, 3).value = desc
            ws_host.cell(rptr, 4).value = risk_letter
        
            res_cell = ws_host.cell(rptr, 5)
            display_res = res
            # 양호인데 리메디 로그가 남아있으면 '조치'로 표시
            # (실제 취약/양호 판단은 res 기준 유지)
            # NOTE: remedy_result는 아래에서 계산하므로 여기서는 빈 값 처리
            #       아래에서 remedy_result 계산 후 다시 display_res 보정
            res_cell.value = display_res
            if display_res == "취약":
                res_cell.fill = vuln_fill
                res_cell.font = vuln_font
            else:
                res_cell.font = normal_font
        
            # 조치결과: 공백
            ws_host.cell(rptr, 6).value = ""
        
            scan_basis = get_scan_basis(hostname, pflag)
            remedy_text = get_remedy_result(hostname, pflag)
            remedy_success = get_remedy_success(hostname, pflag)

            # 진단결과 표시값(취약/조치/양호)
            display_res = res
            if res == "양호" and remedy_success:
                display_res = "조치"

            # res_cell은 위에서 생성됨. 표시값/색상 보정
            res_cell.value = display_res
            if display_res == "취약":
                res_cell.fill = vuln_fill
                res_cell.font = vuln_font
            else:
                res_cell.font = normal_font

            # 엑셀 컬럼 채우기 정책(최종)
            if display_res == "취약":
                ws_host.cell(rptr, 7).value = scan_basis
                ws_host.cell(rptr, 8).value = ""
            elif display_res == "조치":
                ws_host.cell(rptr, 7).value = ""
                ws_host.cell(rptr, 8).value = f"[조치 로그]\n{remedy_text}\n\n[재검사 결과]\n{scan_basis}".strip()




            elif display_res == "양호":
                ws_host.cell(rptr, 7).value = ""
                ws_host.cell(rptr, 8).value = scan_basis
            else:
                ws_host.cell(rptr, 7).value = ""
                ws_host.cell(rptr, 8).value = ""

            # 점수: 공백
            ws_host.cell(rptr, 9).value = ""
        
            for c in range(1, 10):
                cell = ws_host.cell(rptr, c)
                cell.border = border
                cell.alignment = left if c in [3, 7, 8] else center
        
            last_cat = cat
            rptr += 1

        _set_col_widths(
            ws_host,
            {
                1: 18,
                2: 10,
                3: 42,
                4: 8,
                5: 10,
                6: 12,
                7: 28,
                8: 32,
                9: 8,
            },
        )
        ws_host.freeze_panes = f"A{start_row+1}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_unix_report_excel_all_hosts() -> io.BytesIO:
    """
    [호환 유지용]
    기존 UI/호출부에서 사용하던 함수명을 유지하기 위한 wrapper.
    """
    return generate_unix_report_excel()

def _get_client_ip_fallback() -> str:
    """
    Streamlit 환경에서 클라이언트 IP를 최대한 추정한다.
    - 우선순위: X-Forwarded-For -> X-Real-IP -> Remote-Addr
    - Streamlit v1.37+ : st.context.headers
    - 그 이하/기타 : _get_websocket_headers (비공식)
    """
    # 1) Streamlit 1.37+ st.context.headers
    try:
        headers = getattr(st, "context", None)
        if headers is not None and getattr(st.context, "headers", None) is not None:
            h = st.context.headers
            xff = h.get("X-Forwarded-For") or h.get("x-forwarded-for")
            if xff:
                # "client, proxy1, proxy2" 형태일 수 있음
                return xff.split(",")[0].strip()
            xri = h.get("X-Real-IP") or h.get("x-real-ip")
            if xri:
                return str(xri).strip()
            ra = h.get("Remote-Addr") or h.get("remote-addr")
            if ra:
                return str(ra).strip()
    except Exception:
        pass

    # 2) 비공식: websocket headers
    try:
        from streamlit.web.server.websocket_headers import _get_websocket_headers  # type: ignore
        h = _get_websocket_headers() or {}
        xff = h.get("X-Forwarded-For") or h.get("x-forwarded-for")
        if xff:
            return xff.split(",")[0].strip()
        xri = h.get("X-Real-IP") or h.get("x-real-ip")
        if xri:
            return str(xri).strip()
        ra = h.get("Remote-Addr") or h.get("remote-addr")
        if ra:
            return str(ra).strip()
    except Exception:
        pass

    return "unknown"


def write_dashboard_log(user_id: int, action_code: str, details: dict | str | None = None, ip_address: str | None = None) -> None:
    """
    dashboard.logs에 로그 1건 적재
    - details는 dict면 JSON 문자열로 저장
    - ip_address 미지정 시 헤더 기반으로 최대 추정
    """
    if ip_address is None:
        ip_address = _get_client_ip_fallback()

    if details is None:
        details_str = None
    elif isinstance(details, str):
        details_str = details
    else:
        try:
            details_str = json.dumps(details, ensure_ascii=False)
        except Exception:
            details_str = str(details)

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO dashboard.logs (user_id, action_code, details, ip_address)
            VALUES (%s, %s, %s, %s);
            """,
            (user_id, action_code, details_str, ip_address),
        )
        conn.commit()
        conn.close()
    except Exception as e:
        # 로그 실패가 서비스 흐름을 깨면 안 되므로 조용히 넘김(필요 시 print)
        print(f"[dashboard.logs] insert failed: {e}")


def _shrink_details(details: dict | str | None, max_chars: int = 8000) -> dict | str | None:
    """
    details가 과도하게 커져 DB 적재 부담이 생기는 것을 방지.
    - dict면 json dump 후 길이 확인
    - str이면 그대로 길이 확인
    - 초과 시 '...TRUNCATED' 로 자름
    """
    if details is None:
        return None

    if isinstance(details, dict):
        try:
            s = json.dumps(details, ensure_ascii=False)
        except Exception:
            s = str(details)
        if len(s) <= max_chars:
            return details
        # dict를 자르면 JSON이 깨질 수 있으므로 string으로 저장
        return (s[:max_chars] + "\n...TRUNCATED...")
    else:
        s = str(details)
        if len(s) <= max_chars:
            return details
        return s[:max_chars] + "\n...TRUNCATED..."


def get_parent_flag(flag_id: str) -> str:
    parts = str(flag_id).split("_")
    if len(parts) >= 2:
        return f"{parts[0]}_{parts[1]}"
    return str(flag_id)


def run_ansible_simple(playbook_filename, target_host, extra_vars_dict):
    """
    [Native CLI Execution]
    cwd를 설정하여 /home/audit/isms-ansible 경로에서 직접 명령어를 쏘는 느낌으로 구현
    """
    ANSIBLE_DIR = "/home/audit/isms-ansible"
    env = os.environ.copy()
    env["ANSIBLE_ROLES_PATH"] = f"{ANSIBLE_DIR}/roles"
    env["ANSIBLE_HOST_KEY_CHECKING"] = "False"

    cmd_list = [
        "ansible-playbook",
        f"playbooks/{playbook_filename}",
        "-i",
        "inventory/prod",
        "-l",
        target_host,
        "-e",
        json.dumps(extra_vars_dict),
    ]

    try:
        result = subprocess.run(cmd_list, cwd=ANSIBLE_DIR, capture_output=True, text=True, env=env)
        is_success = result.returncode == 0
        combined_log = f"[COMMAND] cd {ANSIBLE_DIR} && {' '.join(cmd_list)}\n\n"
        if result.stdout:
            combined_log += f"[STDOUT]\n{result.stdout}\n"
        if result.stderr:
            combined_log += f"[STDERR]\n{result.stderr}\n"
        return is_success, combined_log.strip(), " ".join(cmd_list)
    except Exception as e:
        return False, str(e), " ".join(cmd_list)


def get_db_connection():
    return psycopg2.connect(
        host="127.0.0.1",
        port="5432",
        database="isms_audit",
        user="audit",
        password="audit123!",
    )


def parse_diagnostic_log(asset_name, flag_id):
    """
    선택된 flag_id에 해당하는 진단 로그를 파싱하여
    [진단 내용] / [진단 근거] / [진단 결과] 형식으로 반환
    """
    log_dir = "/home/audit/isms-ansible/results/scan/log/"
    group_id = get_parent_flag(flag_id)

    # 정확한 파일 우선, 없으면 parent 파일
    search_patterns = [
        f"{log_dir}{asset_name}_{flag_id}.log",
        f"{log_dir}{asset_name}_{group_id}.log",
    ]

    log_file_path = None
    for pattern in search_patterns:
        if os.path.exists(pattern):
            log_file_path = pattern
            break

    if not log_file_path:
        return "로그 데이터를 찾을 수 없습니다."

    contents = []
    basis_list = []
    status_value = None

    try:
        with open(log_file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        for line in lines:
            line = line.strip()

            # json 라인만 처리
            if not line.startswith("{"):
                continue

            try:
                data = json.loads(line)
            except Exception:
                continue

            section = data.get("section", "")
            title = data.get("title", "")
            cmd = data.get("cmd", "")
            basis = data.get("basis", "")
            status = data.get("status", "")

            # 정확한 하위 플래그 기준 매칭
            if f"[{flag_id}]" not in title and f"[{flag_id}]" not in basis:
                continue

            if section == "[진단 내용]" and cmd:
                contents.append(cmd)

            if section == "[진단 결과]" and basis:
                basis_list.append(basis)
                status_value = status

        # 결과 정리
        result_text = ""

        if contents:
            result_text += "[진단 내용]\n"
            for c in contents:
                result_text += f"{c}\n"
            result_text += "\n"

        if basis_list:
            result_text += "[진단 근거]\n"
            for b in basis_list:
                result_text += f"{b}\n"
            result_text += "\n"

        if status_value:
            result_text += "[진단 결과]\n"
            result_text += f"{status_value}\n"

        return result_text.strip() if result_text else "해당 플래그에 대한 로그를 찾을 수 없습니다."

    except Exception as e:
        return f"로그 파싱 중 오류 발생: {str(e)}"


def check_login(login_id, password):
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        query = """
            SELECT u.user_id, u.username, r.role_name
            FROM dashboard.users u
            JOIN dashboard.roles r ON u.role_id = r.role_id
            WHERE u.login_id = %s AND u.password_hash = %s AND u.is_active = TRUE;
        """
        cur.execute(query, (login_id, password))
        result = cur.fetchone()
        conn.close()

        if result:
            user = {"id": result[0], "name": result[1], "role": result[2]}
            # ✅ LOGIN 로깅(성공 시점 1회)
            try:
                write_dashboard_log(
                    user_id=user["id"],
                    action_code="LOGIN",
                    details=_shrink_details(
                        {
                            "login_id": login_id,
                            "result": "success",
                            "role": user["role"],
                            "ts": datetime.now().isoformat(),
                        }
                    ),
                )
            except Exception:
                pass
            return user
        else:
            return None
    except Exception as e:
        print(f"Login Error: {e}")
        return None


@st.cache_data(ttl=5)
def fetch_data(_refresh_key=None):
    """
    [중요]
    DB에는 이력이 누적(INSERT)되므로, 대시보드에는 host+flag 기준 "최신 1건"만 보여주도록 정규화해서 조회한다.
    - 우선순위: round DESC -> scan_at DESC -> scan_id DESC
    """
    conn = get_db_connection()
    query = """
        WITH latest AS (
            SELECT
                s.*,
                ROW_NUMBER() OVER (
                    PARTITION BY s.host_id, s.flag_id
                    ORDER BY s.round DESC, s.scan_at DESC, s.scan_id DESC
                ) AS rn
            FROM audit.scans s
        )
        SELECT
            h.hostname AS asset_name,
            h.ip_address,
            h.os_version AS os_type,
            l.flag_id,
            g.description,
            g.category,
            g.risk_level,
            CASE WHEN l.is_auto = TRUE THEN 'Ready for Remedy' ELSE 'Manual Review' END AS status,
            CASE WHEN l.is_vul = TRUE THEN 1 ELSE 0 END AS is_vul,
            l.scan_at AS scan_date,
            l.scan_id AS id,
            l.msg,
            g.auto_desc,
            g.is_auto AS guide_is_auto,
            l.round
        FROM audit.hosts h
        LEFT JOIN latest l
            ON l.host_id = h.host_id AND l.rn = 1
        LEFT JOIN audit.guides g
            ON l.flag_id = g.flag_id
        WHERE h.hostname IS NOT NULL
        ORDER BY l.scan_at DESC NULLS LAST, h.hostname;
    """
    df = pd.read_sql(query, conn)
    conn.close()

    df["flag_id"] = df["flag_id"].fillna("N/A")
    df["description"] = df["description"].fillna("No Scan Data")
    df["risk_level"] = df["risk_level"].fillna("Low")
    df["guide_is_auto"] = df["guide_is_auto"].fillna(False)

    def map_risk(x):
        if x in ["High", "high", "상"]:
            return "🔴 High"
        if x in ["Med", "med", "중"]:
            return "🟠 Med"
        return "🟡 Low"

    df["risk_display"] = df["risk_level"].apply(map_risk)
    return df


@st.cache_data(ttl=5)
def fetch_report_data(asset_name=None):
    """
    Detailed Results는 항상 최신만 보여준다.
    - 우선순위: round DESC -> scan_at DESC -> scan_id DESC
    """
    conn = get_db_connection()

    query = """
        WITH latest AS (
            SELECT
                s.*,
                ROW_NUMBER() OVER (
                    PARTITION BY s.host_id, s.flag_id
                    ORDER BY s.round DESC, s.scan_at DESC, s.scan_id DESC
                ) AS rn
            FROM audit.scans s
        )
        SELECT
            l.scan_id,
            h.hostname AS host_id,
            l.flag_id,
            g.category,
            g.risk_level,
            g.description,
            g.auto_desc,
            CASE WHEN l.is_vul = TRUE THEN '취약' ELSE '양호' END AS is_vuln,
            TO_CHAR(l.scan_at, 'YYYYMMDDHH24MI') AS timestamp,
            l.round
        FROM latest l
        JOIN audit.hosts h ON h.host_id = l.host_id
        LEFT JOIN audit.guides g ON l.flag_id = g.flag_id
        WHERE l.rn = 1
    """

    params = None
    if asset_name and asset_name != "All Assets":
        query += " AND h.hostname = %s"
        params = (asset_name,)

    df = pd.read_sql(query, conn, params=params) if params else pd.read_sql(query, conn)
    conn.close()

    df["risk_level"] = df["risk_level"].fillna("Low")
    return df


def create_multisheet_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        for host, subdf in df.groupby("host_id"):
            subdf.to_excel(writer, sheet_name=str(host)[:31], index=False)
    output.seek(0)
    return output


def _touch_watchdog_files(asset_name: str, parent_flag: str):
    """
    watchdog 트리거를 위해 (가능하면) 최신 JSON 파일의 mtime을 갱신
    - scan/json
    - remediate/json
    """
    base_dirs = [
        "/home/audit/isms-ansible/results/scan/json/",
        "/home/audit/isms-ansible/results/remediate/json/",
    ]
    for base_dir in base_dirs:
        pattern = os.path.join(base_dir, f"*{asset_name}*{parent_flag}*.json")
        candidates = glob.glob(pattern)
        if candidates:
            found_file = sorted(candidates, key=os.path.getmtime, reverse=True)[0]
            try:
                os.utime(found_file, None)
            except Exception:
                pass


def _poll_db_latest_is_vul(asset_name: str, flag_id: str, parent_flag: str, prev_scan_date, timeout_sec=20, interval_sec=0.5):
    """
    Individual 로직과 동일한 방식의 DB 반영 대기:
    - 우선 (asset, flag_id)로 찾고
    - 없으면 (asset, parent_flag)로 fallback
    - scan_date 변경 or is_vul==0이면 종료
    """
    latest_is_vul = None
    latest_scan_date = None

    max_tries = int(timeout_sec / interval_sec)
    for attempt in range(max_tries):
        try:
            gdf_updated = fetch_data((st.session_state.refresh_flag, f"batch_poll_{attempt}"))
        except Exception:
            time.sleep(interval_sec)
            continue

        row_updated = gdf_updated[
            (gdf_updated["asset_name"] == asset_name) & (gdf_updated["flag_id"] == flag_id)
        ]
        if row_updated.empty:
            row_updated = gdf_updated[
                (gdf_updated["asset_name"] == asset_name) & (gdf_updated["flag_id"] == parent_flag)
            ]

        if not row_updated.empty:
            latest_is_vul = int(row_updated.iloc[0].get("is_vul", 1))
            latest_scan_date = row_updated.iloc[0].get("scan_date")

            try:
                if prev_scan_date is None:
                    break
                if pd.to_datetime(latest_scan_date) != pd.to_datetime(prev_scan_date):
                    break
                if latest_is_vul == 0:
                    break
            except Exception:
                if latest_is_vul == 0:
                    break

        time.sleep(interval_sec)

    return latest_is_vul, latest_scan_date


def filter_auto_desc_by_flag(auto_desc: str, flag_id: str) -> str:
    """
    auto_desc가 여러 플래그 문구를 한 덩어리로 가지고 있는 경우,
    - 공통 안내(헤더/베이스 문구)는 유지
    - '이유:' 이후에 나오는 플래그별 라인만 해당 flag_id 것만 남김
    """
    if auto_desc is None:
        return ""

    text = str(auto_desc).replace("\r\n", "\n").strip()
    if not text or text.lower() == "nan":
        return ""

    lines = text.split("\n")

    # '이유:' 라인 위치 찾기 (없으면 기존 로직처럼 블록 매칭으로 fallback)
    reason_idx = None
    for i, ln in enumerate(lines):
        if ln.strip() in ["이유:", "사유:", "Reason:", "REASON:"]:
            reason_idx = i
            break

    # fallback: 이유 섹션이 없으면, 기존처럼 flag 포함된 블록만 추출
    if reason_idx is None:
        # 빈 줄로 문단 나누기
        blocks = []
        buf = []
        for ln in lines:
            if ln.strip() == "":
                if buf:
                    blocks.append("\n".join(buf).strip())
                    buf = []
            else:
                buf.append(ln)
        if buf:
            blocks.append("\n".join(buf).strip())

        matched = [b for b in blocks if flag_id in b]
        if matched:
            return "\n\n".join(matched).strip()

        matched_lines = [ln for ln in lines if flag_id in ln]
        if matched_lines:
            return "\n".join(matched_lines).strip()

        return text

    header_lines = lines[: reason_idx + 1]
    tail_lines = lines[reason_idx + 1 :]

    filtered_reason_lines = []
    for ln in tail_lines:
        s = ln.strip()
        if not s:
            filtered_reason_lines.append("")
            continue
        if flag_id in s:
            filtered_reason_lines.append(ln)

    cleaned = []
    prev_blank = False
    for ln in filtered_reason_lines:
        if ln.strip() == "":
            if not prev_blank:
                cleaned.append("")
            prev_blank = True
        else:
            cleaned.append(ln)
            prev_blank = False

    matched_exists = any(flag_id in (ln or "") for ln in cleaned)
    if not matched_exists:
        return "\n".join(header_lines).strip()

    out_lines = header_lines[:]
    if out_lines and out_lines[-1].strip() != "":
        out_lines.append("")
    out_lines.extend(cleaned)

    while out_lines and out_lines[-1].strip() == "":
        out_lines.pop()

    return "\n".join(out_lines).strip()


def main():
    # Refresh flag 초기화
    if "refresh_flag" not in st.session_state:
        st.session_state.refresh_flag = False

    if not st.session_state.authenticated:
        _, col2, _ = st.columns([1, 2, 1])
        with col2:
            st.markdown("<h2 style='text-align: center;'>🛡️ Audit System Login</h2>", unsafe_allow_html=True)
            with st.form("login_form"):
                u_id = st.text_input("ID")
                u_pw = st.text_input("Password", type="password")
                if st.form_submit_button("Login", use_container_width=True):
                    user = check_login(u_id, u_pw)
                    if user:
                        st.session_state.authenticated = True
                        st.session_state.user_info = user
                        st.session_state.refresh_flag = not st.session_state.refresh_flag
                        try:
                            fetch_data.clear()
                        except Exception:
                            pass
                        st.rerun()
                    else:
                        st.error("Login Failed")
        return

    df = fetch_data(st.session_state.refresh_flag)
    role = st.session_state.user_info["role"]

    with st.sidebar:
        st.title("🛡️ Audit System")
        st.caption(f"User: {st.session_state.user_info['name']} ({role})")
        st.markdown("---")

        if st.button("대시보드", use_container_width=True, key="nav_dashboard"):
            st.session_state.current_page = "Dashboard"

        if role == "ADMIN":
            st.markdown("<div class='nav-section'>점검</div>", unsafe_allow_html=True)
            with st.container():
                if st.button("자산 점검 시작", use_container_width=True, key="nav_start_audit"):
                    st.session_state.current_page = "Start Audit"
                if st.button("상세 결과", use_container_width=True, key="nav_detailed_result"):
                    st.session_state.current_page = "Detailed Results"

            st.markdown("<div class='nav-section'>조치</div>", unsafe_allow_html=True)
            with st.container():
                if st.button("일괄 조치", use_container_width=True, key="nav_batch_remedy"):
                    st.session_state.current_page = "Batch Remedy"
                if st.button("개별 조치", use_container_width=True, key="nav_individual_remedy"):
                    st.session_state.current_page = "Individual Remedy"

            st.markdown("<div class='nav-section'>수동 검토</div>", unsafe_allow_html=True)
            with st.container():
                if st.button("수동 검토", use_container_width=True, key="nav_manual_review"):
                    st.session_state.current_page = "Manual Review"

            st.markdown("<div class='nav-section'>보고서</div>", unsafe_allow_html=True)
            with st.container():
                if st.button("보고서", use_container_width=True, key="nav_report"):
                    st.session_state.current_page = "Report Export"

        st.markdown("---")

    page = st.session_state.current_page

    # =========================
    # Page display labels (UI only)
    # - current_page 값(라우팅/로직)은 그대로 유지
    # =========================
    PAGE_LABELS = {
        "Start Audit": "자산 점검 시작",
        "Detailed Results": "상세 결과",
        "Report Export": "보고서",
        "Dashboard": "대시보드",
        "Batch Remedy": "일괄 조치",
        "Individual Remedy": "개별 조치",
        "Manual Review": "수동 검토",
    }
    display_page = PAGE_LABELS.get(page, page)

    left, right = st.columns([8, 2])
    with left:
        st.markdown(
            f'<div class="main-header"><h3>{display_page}</h3>'
            f'<div class="header-right"></div></div>',
            unsafe_allow_html=True,
        )
    with right:
        st.markdown(f"<div class='header-user'>👤 {st.session_state.user_info['name']} ({role})</div>", unsafe_allow_html=True)
        if st.button("Logout", use_container_width=True, key="btn_logout_top"):
            st.session_state.authenticated = False
            st.rerun()

    if page == "Dashboard":
        st.markdown("<br>", unsafe_allow_html=True)
        with st.container():
            c_cat, c_opts, c_asset = st.columns([2, 5, 3])
            with c_cat:
                target_category = st.selectbox(
                    "Category",
                    ["Operating System (OS)", "Database (DB)"],
                    label_visibility="collapsed",
                )
            is_os_view = target_category == "Operating System (OS)"
            type_col = "os_type"
            options = sorted(
                df[df[type_col].str.contains("SQL|Postgres", case=False) != is_os_view][type_col]
                .unique()
                .tolist()
            )
            session_key = "os_selection" if is_os_view else "db_selection"
            with c_opts:
                if not st.session_state[session_key] and options:
                    st.session_state[session_key] = [options[0]]
                st.session_state[session_key] = st.pills(
                    f"{target_category}",
                    options,
                    selection_mode="multi",
                    default=st.session_state[session_key],
                    key=f"p_{session_key}",
                    label_visibility="collapsed",
                )
            step1_df = df[df[type_col].isin(st.session_state[session_key])] if st.session_state[session_key] else df[0:0]
            available_assets = sorted(step1_df["asset_name"].unique().tolist())
            with c_asset:
                selected_assets = st.multiselect(
                    "Select Nodes",
                    available_assets,
                    placeholder="All Assets",
                    label_visibility="collapsed",
                )

        f_df = step1_df[step1_df["asset_name"].isin(selected_assets)] if selected_assets else step1_df
        k1, k2, k3, k4 = st.columns(4)

        if f_df.empty:
            total_count = 0
            v_count = 0
            v_rate = 0
            high_risk_count = 0
            auto_count = 0
        else:
            tmp = f_df.copy()
            tmp["parent_flag"] = tmp["flag_id"].astype(str).apply(get_parent_flag)

            def risk_rank(x: str) -> int:
                s = str(x).lower()
                if "high" in s or "상" in s:
                    return 3
                if "med" in s or "중" in s:
                    return 2
                return 1

            grp = tmp.groupby(["asset_name", "parent_flag"], as_index=False).agg(
                is_vul=("is_vul", "max"),
                risk_level=("risk_level", lambda s: sorted(list(s), key=risk_rank, reverse=True)[0] if len(s) else "Low"),
                status=("status", lambda s: "Ready for Remedy" if (s == "Ready for Remedy").any() else "Manual Review"),
            )

            total_count = len(grp)
            v_count = int(grp["is_vul"].sum())
            v_rate = (v_count / total_count * 100) if total_count > 0 else 0

            high_risk_count = len(
                grp[
                    (grp["is_vul"] == 1)
                    & (grp["risk_level"].astype(str).str.contains("High|상", case=False, regex=True))
                ]
            )
            auto_count = len(grp[(grp["is_vul"] == 1) & (grp["status"] == "Ready for Remedy")])

        k1.markdown(
            f'<div class="kpi-card" style="border-left-color:#6C757D;">'
            f'<div class="kpi-title">Assets</div><div class="kpi-value">{f_df["asset_name"].nunique()}</div></div>',
            unsafe_allow_html=True,
        )
        k2.markdown(
            f'<div class="kpi-card" style="border-left-color:#FD7E14;">'
            f'<div class="kpi-title">Vuln Rate</div>'
            f'<div class="kpi-value">{v_rate:.1f}%</div>'
            f'<div style="font-size:11px; opacity:0.75; margin-top:-3px;">({v_count} / {total_count})</div>'
            f'<div style="font-size:10px; opacity:0.6; margin-top:2px;">※ 대분류(U_XX) 기준 집계 · 하위(U_XX_1…)는 상세에서 확인</div>'
            f"</div>",
            unsafe_allow_html=True,
        )
        k3.markdown(
            f'<div class="kpi-card" style="border-left-color:#DC3545;">'
            f'<div class="kpi-title">High Risk</div><div class="kpi-value">{high_risk_count}</div></div>',
            unsafe_allow_html=True,
        )
        k4.markdown(
            f'<div class="kpi-card" style="border-left-color:#20C997;">'
            f'<div class="kpi-title">Auto-Remediable</div><div class="kpi-value">{auto_count}</div></div>',
            unsafe_allow_html=True,
        )

        c1, c2 = st.columns([1, 2])

        vul_rows = f_df[f_df["is_vul"] == 1] if not f_df.empty else f_df

        if (not f_df.empty) and (not vul_rows.empty):
            color_map = {"🔴 High": "#DC3545", "🟠 Med": "#FD7E14", "🟡 Low": "#FFC107"}

            fig_pie = px.pie(
                vul_rows,
                names="risk_display",
                hole=0.6,
                color="risk_display",
                color_discrete_map=color_map,
            )
            fig_pie.update_traces(hovertemplate="<b>%{label}</b><br>Count: %{value}건<br>Ratio: %{percent}")

            with c1:
                st.plotly_chart(fig_pie, use_container_width=True)

            with c2:
                vc = vul_rows["category"].value_counts().reset_index()
                vc.columns = ["category", "count"]

                if vc.empty:
                    st.info("취약 카테고리 집계 데이터가 없습니다.")
                else:
                    fig_bar = px.bar(vc, x="category", y="count")
                    fig_bar.update_layout(
                        height=320,
                        margin=dict(l=10, r=10, t=10, b=10),
                        paper_bgcolor="rgba(0,0,0,0)",
                        plot_bgcolor="rgba(0,0,0,0)",
                    )
                    st.plotly_chart(fig_bar, use_container_width=True)
        else:
            st.info("취약 항목이 없어 그래프를 표시하지 않습니다.")

        st.subheader("Asset Status Summary")

        if not f_df.empty:
            stat = (
                f_df.groupby(["asset_name", "os_type"])
                .agg(Total=("id", "count"), Vulnerable=("is_vul", "sum"))
                .reset_index()
            )
            stat["Vuln Rate (%)"] = (stat["Vulnerable"] / stat["Total"]) * 100

            stat = stat.sort_values(["Vulnerable", "Vuln Rate (%)"], ascending=False).copy()

            stat_disp = stat.copy()
            stat_disp["Total"] = stat_disp["Total"].astype(int).astype(str)
            stat_disp["Vulnerable"] = stat_disp["Vulnerable"].astype(int).astype(str)

            st.dataframe(
                stat_disp,
                hide_index=True,
                use_container_width=True,
                column_config={
                    "asset_name": st.column_config.TextColumn("asset_name", width="medium"),
                    "os_type": st.column_config.TextColumn("os_type", width="medium"),
                    "Total": st.column_config.TextColumn("Total", width="small"),
                    "Vulnerable": st.column_config.TextColumn("Vulnerable", width="small"),
                    "Vuln Rate (%)": st.column_config.ProgressColumn(
                        "Vuln Rate (%)",
                        min_value=0.0,
                        max_value=100.0,
                        format="%.2f",
                        width="medium",
                    ),
                },
            )
        else:
            st.info("표시할 자산 데이터가 없습니다.")

    elif page == "Start Audit" and role == "ADMIN":
        st.markdown("### 자산 점검 시작")
        st.caption("점검할 자산을 선택한 뒤, 아래 버튼으로 점검을 실행하세요.")

        if st.session_state.get("last_audit_summary"):
            s = st.session_state.last_audit_summary
            st.success("✅ 점검 완료")

            c1, c2, c3 = st.columns(3)
            c1.metric("대상 자산", f"{s['total']}개")
            c2.metric("성공/실패", f"{s['ok']} / {s['fail']}")
            c3.metric("소요 시간", f"{s['elapsed_sec']}초")

            if s.get("fail", 0) > 0 and s.get("failed_hosts"):
                with st.expander("❗ 실패한 자산 목록 보기", expanded=False):
                    st.write(", ".join(s["failed_hosts"]))

            if st.button("요약 닫기", key="btn_close_audit_summary"):
                st.session_state.last_audit_summary = None
                st.rerun()

        all_hosts = df[["asset_name", "ip_address", "os_type"]].drop_duplicates().sort_values("asset_name")
        all_hosts.insert(0, "Selected", False)

        selected_data = st.data_editor(
            all_hosts,
            hide_index=True,
            use_container_width=True,
            key="audit_asset_selector",
            column_config={
                "Selected": st.column_config.CheckboxColumn("Selected", width="small"),
                "asset_name": st.column_config.TextColumn("asset_name", width="medium"),
                "ip_address": st.column_config.TextColumn("ip_address", width="medium"),
                "os_type": st.column_config.TextColumn("os_type", width="medium"),
            },
        )
        chosen_hosts = selected_data[selected_data["Selected"]]["asset_name"].tolist()

        if st.button(
            f"{len(chosen_hosts)}개 자산 점검 시작",
            type="secondary",
            use_container_width=True,
            disabled=(len(chosen_hosts) == 0),
        ):
            start_ts = time.time()
            ok_cnt, fail_cnt = 0, 0
            failed_hosts = []

            with st.status("보안 점검 실행 중...", expanded=True) as status:
                for i, host in enumerate(chosen_hosts, 1):
                    status.update(label=f"[{i}/{len(chosen_hosts)}] {host} 점검 중...", state="running")
                    host_os_row = df[df["asset_name"] == host].iloc[0]
                    target_os = "ubuntu" if "ubuntu" in str(host_os_row["os_type"]).lower() else "rocky"

                    success, output, cmd = run_ansible_simple("check.yml", host, {"isms_os": target_os})
                    if success:
                        ok_cnt += 1
                    else:
                        fail_cnt += 1
                        failed_hosts.append(host)

                elapsed = int(time.time() - start_ts)
                status.update(label="점검 완료", state="complete")

            st.session_state.last_audit_summary = {
                "total": len(chosen_hosts),
                "ok": ok_cnt,
                "fail": fail_cnt,
                "failed_hosts": failed_hosts,
                "elapsed_sec": elapsed,
            }

            # ✅ SCAN 로깅(점검 실행 종료 시점 1회)
            try:
                write_dashboard_log(
                    user_id=st.session_state.user_info.get("id", 0),
                    action_code="SCAN",
                    details=_shrink_details(
                        {
                            "hosts": chosen_hosts,
                            "total": len(chosen_hosts),
                            "ok": ok_cnt,
                            "fail": fail_cnt,
                            "failed_hosts": failed_hosts,
                            "elapsed_sec": elapsed,
                            "ts": datetime.now().isoformat(),
                        }
                    ),
                )
            except Exception:
                pass

            st.session_state.refresh_flag = not st.session_state.refresh_flag
            try:
                fetch_data.clear()
            except Exception:
                pass
            st.rerun()

        if st.session_state.execution_output:
            st.markdown("#### Execution Log")
            st.markdown(
                f'<div class="execution-log">{st.session_state.execution_output.replace(chr(10), "<br>")}</div>',
                unsafe_allow_html=True,
            )
            if st.button("로그 창 닫기"):
                st.session_state.execution_output = ""
                st.rerun()

    elif page == "Detailed Results" and role == "ADMIN":
        st.markdown("### 상세 점검 결과")
        st.caption("노드를 선택하면 최신 점검 결과를 확인할 수 있습니다.")

        full_report_data = fetch_report_data("All Assets")
        if not full_report_data.empty:
            all_hosts = sorted(full_report_data["host_id"].unique().tolist())
            selected_host = st.selectbox(" 상세 결과를 확인할 노드를 선택하세요", all_hosts)

            if selected_host != "선택하세요...":
                cols = ["flag_id", "category", "risk_level", "is_vuln", "description", "auto_desc"]
                node_df = full_report_data[full_report_data["host_id"] == selected_host][cols].copy()

                def _map_risk_display(x):
                    s = str(x)
                    if s in ["High", "high", "상"]:
                        return "🔴 High"
                    if s in ["Med", "med", "중"]:
                        return "🟠 Med"
                    return "🟡 Low"

                node_df["risk_display"] = node_df["risk_level"].apply(_map_risk_display)

                # 표시 컬럼 (UI only)
                disp_cols = ["flag_id", "category", "risk_display", "is_vuln", "description", "auto_desc"]
                node_disp = node_df[disp_cols].copy()

                def _style_is_vuln(v):
                    s = str(v)
                    if "취약" in s:
                        return "color:#DC2626;font-weight:700;"
                    return ""

                st.dataframe(
                    node_disp.style.applymap(_style_is_vuln, subset=["is_vuln"]),
                    hide_index=True,
                    use_container_width=True,
                    column_config={
                        "flag_id": st.column_config.TextColumn("flag", width="small"),
                        "category": st.column_config.TextColumn("category", width="medium"),
                        "risk_display": st.column_config.TextColumn("risk", width="small"),
                        "is_vuln": st.column_config.TextColumn("취약 여부", width="small"),
                        "description": st.column_config.TextColumn("description", width="large"),
                        "auto_desc": st.column_config.TextColumn("자동 조치 정보", width="large"),
                    },
                )
        else:
            st.info("표시할 점검 데이터가 없습니다.")

    elif page == "Batch Remedy" and role == "ADMIN":
        st.markdown("### 일괄 조치 (체크된 항목 전체 실행)")

        if st.session_state.get("last_batch_result"):
            br = st.session_state.last_batch_result
            overall_state = br.get("overall_state", "warn")

            if overall_state == "success":
                st.success("✅ 일괄 조치 완료 (모든 항목 정상 반영)")
            elif overall_state == "partial":
                st.warning("⚠️ 일괄 조치 결과 확인 필요 (타임아웃/잔존 취약 포함)")
            else:
                st.error("❌ 일괄 조치 실패 (조치/재점검 단계 오류 포함)")

            c1, c2, c3 = st.columns(3)
            c1.metric("대상 자산", f"{br.get('total_hosts', 0)}개")
            c2.metric("대상 항목", f"{br.get('total_items', 0)}개")
            c3.metric("소요 시간", f"{br.get('elapsed_sec', 0)}초")

            st.markdown(
                f"<div class='result-card'>"
                f"<div class='result-title'> Summary</div>"
                f"<div class='muted'>"
                f"resolved={br.get('resolved_items', 0)} · pending={br.get('pending_items', 0)} · timeout={br.get('timeout_items', 0)}"
                f"</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

            with st.expander("상세 결과 보기", expanded=True):
                for host, hinfo in br.get("hosts", {}).items():
                    host_title = f"{host} — sent_flags={hinfo.get('sent_flags', [])}"
                    st.markdown(f"<div class='result-card'><div class='result-title'>{host_title}</div>", unsafe_allow_html=True)

                    per_rows = []
                    for item in hinfo.get("items", []):
                        per_rows.append(
                            {
                                "flag_id": item.get("flag_id"),
                                "parent_flag": item.get("parent_flag"),
                                "result": item.get("result"),
                                "latest_is_vul": item.get("latest_is_vul"),
                                "note": item.get("note"),
                            }
                        )
                    if per_rows:
                        st.dataframe(pd.DataFrame(per_rows), hide_index=True, use_container_width=True)

                    if hinfo.get("details"):
                        with st.expander(f"로그 보기 ({host})", expanded=False):
                            st.code(hinfo["details"], language="text")

                    st.markdown("</div>", unsafe_allow_html=True)

            if st.button("결과창 닫기", key="btn_close_batch_result"):
                st.session_state.last_batch_result = None
                st.rerun()

        gdf = df[(df["is_vul"] == 1) & (df["guide_is_auto"] == True)].copy()

        if gdf.empty:
            st.success("현재 자동 조치 가능한 취약 항목이 없습니다.")
        else:
            if "Selected" not in gdf.columns:
                gdf.insert(0, "Selected", False)

            top1, top2 = st.columns([8, 2])
            with top1:
                batch_q = st.text_input(
                    "Search",
                    key="batch_search",
                    placeholder="asset / flag / description / category 로 검색",
                    label_visibility="collapsed",
                )
            with top2:
                batch_select_all = st.checkbox("전체 선택", key="batch_select_all")

            view_df = gdf.copy()
            if batch_q:
                q = batch_q.strip().lower()
                view_df = view_df[
                    view_df["asset_name"].astype(str).str.lower().str.contains(q)
                    | view_df["flag_id"].astype(str).str.lower().str.contains(q)
                    | view_df["description"].astype(str).str.lower().str.contains(q)
                    | view_df["category"].astype(str).str.lower().str.contains(q)
                ]

            if batch_select_all and not view_df.empty:
                view_df.loc[:, "Selected"] = True

            grid_cols = ["Selected", "asset_name", "os_type", "flag_id", "risk_display", "category", "description"]
            view_df = view_df[grid_cols].copy()

            edited_df = st.data_editor(
                view_df,
                hide_index=True,
                use_container_width=True,
                column_config={
                    "Selected": st.column_config.CheckboxColumn(
                        "Selected",
                        help="조치할 항목을 선택하세요",
                        default=False,
                        width="small",
                    ),
                    "asset_name": st.column_config.TextColumn("asset_name", width="medium"),
                    "os_type": st.column_config.TextColumn("os_type", width="small"),
                    "flag_id": st.column_config.TextColumn("flag_id", width="small"),
                    "risk_display": st.column_config.TextColumn("risk_display", width="small"),
                    "category": st.column_config.TextColumn("category", width="small"),
                    "description": st.column_config.TextColumn("description", width="large"),
                },
                key="batch_editor",
            )

            selected_cnt = int(edited_df["Selected"].sum()) if "Selected" in edited_df.columns else 0

            if st.button(
                f"선택된 {selected_cnt}개 항목 일괄 조치 실행",
                type="secondary",
                use_container_width=True,
                disabled=(selected_cnt == 0),
            ):
                start_ts = time.time()

                selected_items = edited_df[edited_df["Selected"]].copy()

                host_groups = {}
                host_items = {}
                for _, row in selected_items.iterrows():
                    h = row["asset_name"]
                    f = str(row["flag_id"])
                    p = get_parent_flag(f)

                    if h not in host_groups:
                        host_groups[h] = []
                        host_items[h] = []

                    for cand in [p, f]:
                        if cand not in host_groups[h]:
                            host_groups[h].append(cand)

                    host_items[h].append({"flag_id": f, "parent_flag": p})

                batch_result = {
                    "overall_state": "warn",
                    "total_hosts": len(host_groups),
                    "total_items": int(selected_cnt),
                    "resolved_items": 0,
                    "pending_items": 0,
                    "timeout_items": 0,
                    "elapsed_sec": 0,
                    "hosts": {},
                }

                any_fail = False
                any_partial = False

                with st.status("일괄 조치 진행 중...", expanded=True) as status:
                    for hi, (host, flags) in enumerate(host_groups.items(), 1):
                        status.update(label=f"[{hi}/{len(host_groups)}] {host} 조치 실행 중...", state="running")

                        host_os_row = df[df["asset_name"] == host].iloc[0]
                        target_os = "ubuntu" if "ubuntu" in str(host_os_row["os_type"]).lower() else "rocky"

                        rem_success, rem_output, _ = run_ansible_simple(
                            "remediate.yml",
                            host,
                            {"flag_list": flags, "isms_os": target_os},
                        )

                        scan_success = False
                        scan_output = ""
                        if rem_success:
                            status.update(label=f"[{hi}/{len(host_groups)}] {host} 재점검 실행 중...", state="running")
                            parent_flags = sorted({it["parent_flag"] for it in host_items.get(host, [])})

                            scan_success, scan_output, _ = run_ansible_simple(
                                "check.yml",
                                host,
                                {"flag_list": parent_flags, "isms_os": target_os},
                            )
                        else:
                            any_fail = True

                        for it in host_items.get(host, []):
                            _touch_watchdog_files(host, it["parent_flag"])

                        items_out = []
                        details_parts = []
                        details_parts.append(f"=== {host} REMEDIATE LOG ===\n{rem_output}")
                        if rem_success:
                            details_parts.append(f"=== {host} RESCAN LOG ===\n{scan_output}")

                        gdf_before = df[df["asset_name"] == host].copy()

                        for it in host_items.get(host, []):
                            f = it["flag_id"]
                            p = it["parent_flag"]

                            prev_scan_date = None
                            prev_row = gdf_before[gdf_before["flag_id"] == f]
                            if prev_row.empty:
                                prev_row = gdf_before[gdf_before["flag_id"] == p]
                            if not prev_row.empty:
                                prev_scan_date = prev_row.iloc[0].get("scan_date")

                            if not rem_success:
                                items_out.append(
                                    {
                                        "flag_id": f,
                                        "parent_flag": p,
                                        "latest_is_vul": None,
                                        "result": "❌ 조치 실패",
                                        "note": "remediate 단계 실패",
                                    }
                                )
                                batch_result["pending_items"] += 1
                                any_fail = True
                                continue

                            if not scan_success:
                                items_out.append(
                                    {
                                        "flag_id": f,
                                        "parent_flag": p,
                                        "latest_is_vul": None,
                                        "result": "❌ 재점검 실패",
                                        "note": "check 단계 실패",
                                    }
                                )
                                batch_result["pending_items"] += 1
                                any_fail = True
                                continue

                            latest_is_vul, latest_scan_date = _poll_db_latest_is_vul(
                                asset_name=host,
                                flag_id=f,
                                parent_flag=p,
                                prev_scan_date=prev_scan_date,
                                timeout_sec=20,
                                interval_sec=0.5,
                            )

                            if latest_is_vul is None:
                                items_out.append(
                                    {
                                        "flag_id": f,
                                        "parent_flag": p,
                                        "latest_is_vul": None,
                                        "result": "⏱️ 타임아웃",
                                        "note": "DB 반영 확인 실패",
                                    }
                                )
                                batch_result["timeout_items"] += 1
                                batch_result["pending_items"] += 1
                                any_partial = True
                            elif int(latest_is_vul) == 0:
                                items_out.append(
                                    {
                                        "flag_id": f,
                                        "parent_flag": p,
                                        "latest_is_vul": int(latest_is_vul),
                                        "result": "✅ 조치 성공",
                                        "note": "취약 제거 확인",
                                    }
                                )
                                batch_result["resolved_items"] += 1
                            else:
                                items_out.append(
                                    {
                                        "flag_id": f,
                                        "parent_flag": p,
                                        "latest_is_vul": int(latest_is_vul),
                                        "result": "⚠️ 잔존 취약",
                                        "note": "조치 후에도 취약 상태",
                                    }
                                )
                                batch_result["pending_items"] += 1
                                any_partial = True

                        details_text = "\n\n".join(details_parts)

                        batch_result["hosts"][host] = {
                            "sent_flags": flags,
                            "remediate_success": bool(rem_success),
                            "rescan_success": bool(scan_success),
                            "items": items_out,
                            "details": details_text,
                        }

                    elapsed = int(time.time() - start_ts)
                    batch_result["elapsed_sec"] = elapsed

                    if any_fail:
                        batch_result["overall_state"] = "fail"
                        status.update(label="일괄 조치 종료 (실패 포함)", state="error")
                    elif any_partial:
                        batch_result["overall_state"] = "partial"
                        status.update(label="일괄 조치 종료 (확인 필요)", state="complete")
                    else:
                        batch_result["overall_state"] = "success"
                        status.update(label="일괄 조치 완료", state="complete")

                st.session_state.last_batch_result = batch_result

                try:
                    summary_for_log = {
                        "overall_state": batch_result.get("overall_state"),
                        "total_hosts": batch_result.get("total_hosts"),
                        "total_items": batch_result.get("total_items"),
                        "resolved_items": batch_result.get("resolved_items"),
                        "pending_items": batch_result.get("pending_items"),
                        "timeout_items": batch_result.get("timeout_items"),
                        "elapsed_sec": batch_result.get("elapsed_sec"),
                        "hosts": {
                            h: {
                                "sent_flags": info.get("sent_flags", []),
                                "remediate_success": info.get("remediate_success"),
                                "rescan_success": info.get("rescan_success"),
                                "items": info.get("items", []),
                            }
                            for h, info in batch_result.get("hosts", {}).items()
                        },
                        "ts": datetime.now().isoformat(),
                    }
                    write_dashboard_log(
                        user_id=st.session_state.user_info.get("id", 0),
                        action_code="REMEDY",
                        details=_shrink_details(summary_for_log),
                    )
                except Exception:
                    pass

                st.session_state.refresh_flag = not st.session_state.get("refresh_flag", False)
                try:
                    fetch_data.clear()
                except Exception:
                    pass
                st.rerun()

    elif page == "Individual Remedy" and role == "ADMIN":
        st.markdown("### 세부 취약점 분석 및 개별 조치")
        i_q = st.text_input(
            "Likely Search",
            key="ind_search",
            placeholder="asset / flag / description / category 로 검색",
            label_visibility="collapsed",
        )

        gdf = fetch_data(st.session_state.refresh_flag)
        gdf = gdf[(gdf["is_vul"] == 1) & (gdf["guide_is_auto"] == True)].copy()
        if i_q:
            q = i_q.strip().lower()
            gdf = gdf[
                gdf["asset_name"].astype(str).str.lower().str.contains(q)
                | gdf["flag_id"].astype(str).str.lower().str.contains(q)
                | gdf["description"].astype(str).str.lower().str.contains(q)
                | gdf["category"].astype(str).str.lower().str.contains(q)
            ]

        if not gdf.empty:
            c_l, c_r = st.columns([2, 3])
            with c_l:
                sel = st.dataframe(
                    gdf[["asset_name", "os_type", "flag_id", "risk_display"]],
                    hide_index=True,
                    use_container_width=True,
                    on_select="rerun",
                    selection_mode="single-row",
                    key="i_sel",
                )

            with c_r:
                if sel.selection.rows:
                    row = gdf.iloc[sel.selection.rows[0]]
                    st.markdown(f"#### asset: `{row['asset_name']}` | flag: `{row['flag_id']}`")
                    st.markdown("**1️⃣ 진단 상세 근거 (Evidence)**")
                    st.code(parse_diagnostic_log(row["asset_name"], row["flag_id"]), language="text")

                    st.markdown("**2️⃣ 자동 조치 안내 (권장 조치)**")
                    auto_desc_text = filter_auto_desc_by_flag(row.get("auto_desc"), row["flag_id"])

                    if auto_desc_text and str(auto_desc_text).strip():
                        st.info(auto_desc_text)
                    else:
                        st.info("등록된 자동 조치 안내가 없습니다.")

                    if st.button("정밀 조치 및 재점검 실행", use_container_width=True, type="secondary"):
                        parent_f = get_parent_flag(row["flag_id"])
                        target_os = "ubuntu" if "ubuntu" in str(row["os_type"]).lower() else "rocky"

                        with st.status("조치 진행 중...", expanded=True) as status:
                            status.update(label=f"{parent_f} 조치 실행 중...", state="running")

                            rem_success, rem_output, _ = run_ansible_simple(
                                "remediate.yml",
                                row["asset_name"],
                                {"flag_list": [parent_f, row["flag_id"]], "isms_os": target_os},
                            )

                            if rem_success:
                                status.update(label=f"{parent_f} 재점검(개별 진단) 중...", state="running")
                                scan_success, scan_output, _ = run_ansible_simple(
                                    "check.yml",
                                    row["asset_name"],
                                    {"flag_list": [parent_f], "isms_os": target_os},
                                )

                                _touch_watchdog_files(row["asset_name"], parent_f)

                                prev_scan_date = row.get("scan_date")
                                latest_is_vul = None

                                latest_is_vul, _latest_scan_date = _poll_db_latest_is_vul(
                                    asset_name=row["asset_name"],
                                    flag_id=row["flag_id"],
                                    parent_flag=parent_f,
                                    prev_scan_date=prev_scan_date,
                                    timeout_sec=15,
                                    interval_sec=0.5,
                                )

                                details_parts = []
                                details_parts.append("=== REMEDIATE LOG ===\n" + rem_output)
                                details_parts.append("=== RESCAN LOG ===\n" + scan_output)
                                details_parts.append(f"=== DB POLL RESULT ===\nlatest_is_vul={latest_is_vul}\nprev_scan_date={prev_scan_date}")
                                details_text = "\n\n".join(details_parts)

                                if scan_success is False:
                                    status.update(label="❌ 재점검 단계 실패", state="error")
                                    st.session_state.last_remedy_result = {
                                        "success": False,
                                        "asset_name": row.get("asset_name"),
                                        "flag_id": row.get("flag_id"),
                                        "details": details_text,
                                    }
                                elif latest_is_vul is None:
                                    status.update(label="⚠️ DB 반영 확인 실패 (타임아웃)", state="error")
                                    st.session_state.last_remedy_result = {
                                        "success": False,
                                        "asset_name": row.get("asset_name"),
                                        "flag_id": row.get("flag_id"),
                                        "details": details_text,
                                    }
                                elif int(latest_is_vul) == 0:
                                    status.update(label="✅ 취약점 제거 완료!", state="complete")
                                    st.session_state.last_remedy_result = {
                                        "success": True,
                                        "asset_name": row.get("asset_name"),
                                        "flag_id": row.get("flag_id"),
                                        "details": details_text,
                                    }
                                else:
                                    status.update(label="⚠️ 조치 후 취약점 여전히 존재", state="error")
                                    st.session_state.last_remedy_result = {
                                        "success": False,
                                        "asset_name": row.get("asset_name"),
                                        "flag_id": row.get("flag_id"),
                                        "details": details_text,
                                    }

                                try:
                                    write_dashboard_log(
                                        user_id=st.session_state.user_info.get("id", 0),
                                        action_code="REMEDY",
                                        details=_shrink_details(
                                            {
                                                "mode": "individual",
                                                "asset_name": row.get("asset_name"),
                                                "flag_id": row.get("flag_id"),
                                                "parent_flag": parent_f,
                                                "isms_os": target_os,
                                                "remediate_success": bool(rem_success),
                                                "rescan_success": bool(scan_success),
                                                "latest_is_vul": latest_is_vul,
                                                "prev_scan_date": str(prev_scan_date),
                                                "ts": datetime.now().isoformat(),
                                            }
                                        ),
                                    )
                                except Exception:
                                    pass

                                st.session_state.refresh_flag = not st.session_state.get("refresh_flag", False)
                                try:
                                    fetch_data.clear()
                                except Exception:
                                    pass
                                st.session_state.needs_refresh_after_remedy = True

                            else:
                                status.update(label="❌ 조치 단계 실패", state="error")
                                st.session_state.last_remedy_result = {
                                    "success": False,
                                    "asset_name": row.get("asset_name"),
                                    "flag_id": row.get("flag_id"),
                                    "details": rem_output,
                                }

                                try:
                                    write_dashboard_log(
                                        user_id=st.session_state.user_info.get("id", 0),
                                        action_code="REMEDY",
                                        details=_shrink_details(
                                            {
                                                "mode": "individual",
                                                "asset_name": row.get("asset_name"),
                                                "flag_id": row.get("flag_id"),
                                                "parent_flag": get_parent_flag(row.get("flag_id")),
                                                "result": "remediate_failed",
                                                "ts": datetime.now().isoformat(),
                                            }
                                        ),
                                    )
                                except Exception:
                                    pass

                                st.session_state.refresh_flag = not st.session_state.get("refresh_flag", False)
                                try:
                                    fetch_data.clear()
                                except Exception:
                                    pass
                                st.session_state.needs_refresh_after_remedy = True

                    _res = st.session_state.last_remedy_result
                    if _res and _res.get("asset_name") == row.get("asset_name") and _res.get("flag_id") == row.get("flag_id"):
                        if _res.get("success") is True:
                            st.success("✅ 조치 성공")
                        else:
                            st.error("⚠️ 조치 실패")
                            details = _res.get("details")
                            if details:
                                with st.expander("실패 상세 로그", expanded=True):
                                    st.code(details, language="text")

                        if st.button("결과 닫기", key="close_inline_result"):
                            st.session_state.last_remedy_result = None
                            st.session_state.needs_refresh_after_remedy = False
                            st.rerun()

                    if st.session_state.needs_refresh_after_remedy:
                        if st.button("🔄 목록 새로고침", key="btn_refresh_after_remedy", use_container_width=True):
                            st.session_state.needs_refresh_after_remedy = False
                            st.rerun()
                else:
                    st.info("목록에서 항목을 선택하여 상세 정보와 조치 기능을 확인하세요.")
        else:
            st.success("현재 조치가 필요한 취약한 항목이 없습니다.")

    elif page == "Manual Review":
        st.markdown("### 수동 검토 대기열 (취약 & 자동조치 불가)")

        m_q = st.text_input(
            "Likely Search",
            key="manual_search",
            placeholder="asset / flag / description / category 로 검색",
            label_visibility="collapsed",
        )

        mdf = df[(df["is_vul"] == 1) & (df["guide_is_auto"] == False)].copy()

        if m_q:
            q = m_q.strip().lower()
            mdf = mdf[
                mdf["asset_name"].astype(str).str.lower().str.contains(q)
                | mdf["flag_id"].astype(str).str.lower().str.contains(q)
                | mdf["description"].astype(str).str.lower().str.contains(q)
                | mdf["category"].astype(str).str.lower().str.contains(q)
            ]

        if not mdf.empty:
            mdf["자동 조치 불가 사유"] = mdf.apply(
                lambda r: filter_auto_desc_by_flag(r.get("auto_desc"), r.get("flag_id")),
                axis=1,
            )

        if mdf.empty:
            st.success("현재 수동 검토 대상(취약 & 자동조치 불가) 항목이 없습니다.")
        else:
            show_cols = ["asset_name", "os_type", "flag_id", "category", "description", "자동 조치 불가 사유"]
            st.dataframe(mdf[show_cols], hide_index=True, use_container_width=True)

    elif page == "Report Export" and role == "ADMIN":
        st.markdown("### 보고서 내보내기 (UNIX)")

        st.caption("현재 등록된 전체 노드 중 UNIX 키워드에 매칭되는 노드만 포함하여 보고서를 생성합니다.")

        if st.button("UNIX 보고서 생성", type="secondary", use_container_width=True):
            report_bytes = generate_unix_report_excel_all_hosts()

            fname = f"서버_취약점진단_상세결과_UNIX_{datetime.now().strftime('%Y%m%d')}.xlsx"
            st.download_button("다운로드", report_bytes, fname)

            try:
                write_dashboard_log(
                    user_id=st.session_state.user_info.get("id", 0),
                    action_code="REPORT",
                    details=_shrink_details(
                        {
                            "report_type": "UNIX_CUSTOM",
                            "unix_keywords": UNIX_KEYWORDS,
                            "filename": fname,
                            "ts": datetime.now().isoformat(),
                        }
                    ),
                )
            except Exception:
                pass


if __name__ == "__main__":
    main()